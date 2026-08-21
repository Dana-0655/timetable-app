
import os
import uuid
import threading
import io
import pandas as pd
from flask_cors import CORS
# pyrefly: ignore [missing-import]
from dotenv import load_dotenv
# pyrefly: ignore [missing-import]
from flask import Flask, request, jsonify, send_file
# pyrefly: ignore [missing-import]
from flask_sqlalchemy import SQLAlchemy
# pyrefly: ignore [missing-import]
from flask_bcrypt import Bcrypt
# pyrefly: ignore [missing-import]
from apscheduler.schedulers.background import BackgroundScheduler
from datetime import datetime, timedelta
# pyrefly: ignore [missing-import]
from apscheduler.triggers.cron import CronTrigger  
# pyrefly: ignore [missing-import]
from openpyxl import Workbook
from openpyxl.styles import Font, PatternFill, Alignment, Border, Side
# pyrefly: ignore [missing-import]
from ortools.sat.python import cp_model
from openpyxl.worksheet.datavalidation import DataValidation
from solver import generate_timetable_with_ortools # Import your solver function

load_dotenv()

app = Flask(__name__)
CORS(app)

app.config['SQLALCHEMY_DATABASE_URI'] = os.environ.get('DATABASE_URL')

app.config["SQLALCHEMY_ENGINE_OPTIONS"] = {
    "connect_args": {"ssl": {"ca": "ca.pem"}},
    "pool_pre_ping": True,
    "pool_recycle": 280,
}
app.config['SQLALCHEMY_TRACK_MODIFICATIONS'] = False

db = SQLAlchemy(app)
bcrypt = Bcrypt(app)

class College(db.Model):
    college_id = db.Column(db.Integer, primary_key=True)
    college_name = db.Column(db.String(150), nullable=False)
    college_code = db.Column(db.String(20), unique=True, nullable=False)
    created_at = db.Column(db.DateTime, server_default=db.func.now())

class Admin(db.Model):
    admin_id = db.Column(db.Integer, primary_key=True)
    college_id = db.Column(db.Integer, db.ForeignKey('college.college_id'), nullable=False)
    name = db.Column(db.String(100), nullable=False)
    email = db.Column(db.String(150), nullable=False)
    password_hash = db.Column(db.String(255), nullable=False)
    department_name = db.Column(db.String(100), nullable=False)
    created_at = db.Column(db.DateTime, server_default=db.func.now())

    __table_args__ = (
        db.UniqueConstraint('college_id', 'email', name='unique_admin_email_per_college'),
    )

class Faculty(db.Model):
    faculty_id = db.Column(db.Integer, primary_key=True)
    college_id = db.Column(db.Integer, db.ForeignKey('college.college_id'), nullable=False)
    name = db.Column(db.String(100), nullable=False)
    email = db.Column(db.String(150), nullable=False)
    password_hash = db.Column(db.String(255), nullable=False)
    subject_expertise = db.Column(db.String(255), nullable=True)
    created_at = db.Column(db.DateTime, server_default=db.func.now())

    __table_args__ = (
        db.UniqueConstraint('college_id', 'email', name='unique_faculty_email_per_college'),
    )
class Class(db.Model):
    class_id = db.Column(db.Integer, primary_key=True)
    college_id = db.Column(db.Integer, db.ForeignKey('college.college_id'), nullable=False)
    semester_id = db.Column(db.Integer, db.ForeignKey('semester.semester_id'), nullable=True)
    year = db.Column(db.String(20), nullable=False)
    section = db.Column(db.String(10), nullable=False)
    department = db.Column(db.String(100), nullable=False)
    cc_faculty_id = db.Column(db.Integer, db.ForeignKey('faculty.faculty_id'), nullable=True)
    room_number = db.Column(db.String(50), nullable=True)
    created_at = db.Column(db.DateTime, server_default=db.func.now())

class CCRequest(db.Model):
    cc_request_id = db.Column(db.Integer, primary_key=True)
    class_id = db.Column(db.Integer, db.ForeignKey('class.class_id'), nullable=False)
    faculty_id = db.Column(db.Integer, db.ForeignKey('faculty.faculty_id'), nullable=False)
    initiated_by = db.Column(db.String(20), nullable=False)  # 'faculty' or 'admin'
    status = db.Column(db.String(20), nullable=False, default='pending')  # pending/accepted/rejected
    requested_at = db.Column(db.DateTime, server_default=db.func.now())
    resolved_at = db.Column(db.DateTime, nullable=True)

class Course(db.Model):
    course_id = db.Column(db.Integer, primary_key=True)
    class_id = db.Column(db.Integer, db.ForeignKey('class.class_id'), nullable=False)
    course_name = db.Column(db.String(100), nullable=False)
    course_code = db.Column(db.String(30), nullable=True)
    faculty_id = db.Column(db.Integer, db.ForeignKey('faculty.faculty_id'), nullable=True)
    created_at = db.Column(db.DateTime, server_default=db.func.now())

    __table_args__ = (
        db.UniqueConstraint('class_id', 'course_name', name='unique_course_per_class'),
    )

class CourseFacultyRequest(db.Model):
    cf_request_id = db.Column(db.Integer, primary_key=True)
    course_id = db.Column(db.Integer, db.ForeignKey('course.course_id'), nullable=False)
    faculty_id = db.Column(db.Integer, db.ForeignKey('faculty.faculty_id'), nullable=False)
    initiated_by = db.Column(db.String(20), nullable=False)  # 'faculty', 'cc', or 'admin'
    status = db.Column(db.String(20), nullable=False, default='pending')
    requested_at = db.Column(db.DateTime, server_default=db.func.now())
    resolved_at = db.Column(db.DateTime, nullable=True)

class TimetableConfig(db.Model):
    config_id = db.Column(db.Integer, primary_key=True)
    class_id = db.Column(db.Integer, db.ForeignKey('class.class_id'), nullable=False)
    periods_per_day = db.Column(db.Integer, nullable=False)
    period_duration_minutes = db.Column(db.Integer, nullable=False)
    start_time = db.Column(db.String(10), nullable=False)  # e.g. "09:00"

class TimetableEntry(db.Model):
    entry_id = db.Column(db.Integer, primary_key=True)
    class_id = db.Column(db.Integer, db.ForeignKey('class.class_id'), nullable=False)
    day_of_week = db.Column(db.String(10), nullable=False)  # MON, TUE, etc.
    period_no = db.Column(db.Integer, nullable=False)
    entry_type = db.Column(db.String(10), nullable=False, default='period')  # 'period' or 'break'
    label = db.Column(db.String(50), nullable=True)  # break name, e.g. "Lunch"
    start_time = db.Column(db.String(10), nullable=True)  # e.g. "09:00"
    end_time = db.Column(db.String(10), nullable=True)  # e.g. "09:50"
    course_id = db.Column(db.Integer, db.ForeignKey('course.course_id'), nullable=True)
    status_color = db.Column(db.String(20), nullable=False, default='normal')

class Leave(db.Model):
    leave_id = db.Column(db.Integer, primary_key=True)
    faculty_id = db.Column(db.Integer, db.ForeignKey('faculty.faculty_id'), nullable=False)
    entry_id = db.Column(db.Integer, db.ForeignKey('timetable_entry.entry_id'), nullable=False)
    leave_date = db.Column(db.String(20), nullable=False)  # e.g. "2026-07-20"
    status = db.Column(db.String(20), nullable=False, default='open')  # open/pending_requests/confirmed
    confirmed_faculty_id = db.Column(db.Integer, db.ForeignKey('faculty.faculty_id'), nullable=True)
    confirmed_by_role = db.Column(db.String(20), nullable=True)  # 'faculty' or 'cc'
    created_at = db.Column(db.DateTime, server_default=db.func.now())

class Holiday(db.Model):
    holiday_id = db.Column(db.Integer, primary_key=True)
    college_id = db.Column(db.Integer, db.ForeignKey('college.college_id'), nullable=False)
    holiday_date = db.Column(db.String(20), nullable=False)  # "YYYY-MM-DD"
    reason = db.Column(db.String(150), nullable=True)
    created_at = db.Column(db.DateTime, server_default=db.func.now())

    __table_args__ = (
        db.UniqueConstraint('college_id', 'holiday_date', name='unique_holiday_per_date'),
    )

class CoverRequest(db.Model):
    cover_req_id = db.Column(db.Integer, primary_key=True)
    leave_id = db.Column(db.Integer, db.ForeignKey('leave.leave_id'), nullable=False)
    requesting_faculty_id = db.Column(db.Integer, db.ForeignKey('faculty.faculty_id'), nullable=False)
    status = db.Column(db.String(20), nullable=False, default='pending')
    requested_at = db.Column(db.DateTime, server_default=db.func.now())

class SwapRequest(db.Model):
    swap_id = db.Column(db.Integer, primary_key=True)
    requester_faculty_id = db.Column(db.Integer, db.ForeignKey('faculty.faculty_id'), nullable=False)
    requester_entry_id = db.Column(db.Integer, db.ForeignKey('timetable_entry.entry_id'), nullable=False)
    target_faculty_id = db.Column(db.Integer, db.ForeignKey('faculty.faculty_id'), nullable=False)
    target_entry_id = db.Column(db.Integer, db.ForeignKey('timetable_entry.entry_id'), nullable=False)
    status = db.Column(db.String(20), nullable=False, default='pending')
    rejection_reason = db.Column(db.String(255), nullable=True)
    requested_at = db.Column(db.DateTime, server_default=db.func.now())
    resolved_at = db.Column(db.DateTime, nullable=True)

class Notification(db.Model):
    notification_id = db.Column(db.Integer, primary_key=True)
    leave_id = db.Column(db.Integer, db.ForeignKey('leave.leave_id'), nullable=False)
    sent_to_faculty = db.Column(db.Boolean, default=False)
    sent_to_cc = db.Column(db.Boolean, default=False)
    created_at = db.Column(db.DateTime, server_default=db.func.now())

class Semester(db.Model):
    semester_id = db.Column(db.Integer, primary_key=True)
    college_id = db.Column(db.Integer, db.ForeignKey('college.college_id'), nullable=False)
    semester_name = db.Column(db.String(50), nullable=False)
    is_active = db.Column(db.Boolean, default=False)
    is_deleted = db.Column(db.Boolean, default=False)
    created_at = db.Column(db.DateTime, server_default=db.func.now())

class UserNotification(db.Model):
    notification_id = db.Column(db.Integer, primary_key=True)
    recipient_type = db.Column(db.String(20), nullable=False)
    recipient_id = db.Column(db.Integer, nullable=False)
    message = db.Column(db.String(255), nullable=False)
    notif_type = db.Column(db.String(30), nullable=False)
    reference_id = db.Column(db.Integer, nullable=True)  # NEW: cc_request_id / cf_request_id / swap_id
    is_read = db.Column(db.Boolean, default=False)
    created_at = db.Column(db.DateTime, server_default=db.func.now())

# Model for tracking async background generation jobs
class Job(db.Model):
    __tablename__ = 'job'
    job_id = db.Column(db.String(50), primary_key=True)
    college_id = db.Column(db.Integer, nullable=True)
    job_type = db.Column(db.String(50), nullable=True)
    status = db.Column(db.String(20), default='pending')  # 'pending', 'running', 'success', 'failed'
    progress_message = db.Column(db.String(255), nullable=True)
    result = db.Column(db.JSON, nullable=True)
    error = db.Column(db.Text, nullable=True)

# Model for classroom / lab facilities
class Room(db.Model):
    __tablename__ = 'room'
    room_id = db.Column(db.Integer, primary_key=True)
    college_id = db.Column(db.Integer, db.ForeignKey('college.college_id'), nullable=False)
    room_name = db.Column(db.String(100), nullable=False)
    room_type = db.Column(db.String(50), default='Lecture')  # 'Lecture', 'Lab', etc.
    capacity = db.Column(db.Integer, nullable=True)

def delete_timetable_entries_cascade(class_ids=None, entry_ids=None):
    """
    Safely delete TimetableEntry records and their dependent rows
    (SwapRequest, CoverRequest, Notification, Leave) to avoid foreign key violations.
    """
    if class_ids is not None:
        if not class_ids:
            return
        entries = TimetableEntry.query.filter(TimetableEntry.class_id.in_(class_ids)).all()
        target_entry_ids = [e.entry_id for e in entries]
    elif entry_ids is not None:
        target_entry_ids = list(entry_ids)
    else:
        return

    if not target_entry_ids:
        return

    # 1. Delete dependent SwapRequests where entry is requester or target
    SwapRequest.query.filter(
        (SwapRequest.requester_entry_id.in_(target_entry_ids)) |
        (SwapRequest.target_entry_id.in_(target_entry_ids))
    ).delete(synchronize_session=False)

    # 2. Find leaves linked to these entries
    leaves = Leave.query.filter(Leave.entry_id.in_(target_entry_ids)).all()
    leave_ids = [l.leave_id for l in leaves]
    if leave_ids:
        # Delete dependent CoverRequests & Notifications
        CoverRequest.query.filter(CoverRequest.leave_id.in_(leave_ids)).delete(synchronize_session=False)
        Notification.query.filter(Notification.leave_id.in_(leave_ids)).delete(synchronize_session=False)
        Leave.query.filter(Leave.leave_id.in_(leave_ids)).delete(synchronize_session=False)

    # 3. Delete the timetable entries themselves
    TimetableEntry.query.filter(TimetableEntry.entry_id.in_(target_entry_ids)).delete(synchronize_session=False)

def reset_weekly_swaps():
    with app.app_context():
        active_swaps = SwapRequest.query.filter_by(status="accepted").all()
        for swap in active_swaps:
            requester_entry = TimetableEntry.query.get(swap.requester_entry_id)
            target_entry = TimetableEntry.query.get(swap.target_entry_id)
            if requester_entry and target_entry:
                requester_entry.course_id, target_entry.course_id = (
                    target_entry.course_id,
                    requester_entry.course_id,
                )
                requester_entry.status_color = "normal"
                target_entry.status_color = "normal"
            db.session.delete(swap)

        db.session.commit()
        print(f"[WEEKLY RESET] Reverted {len(active_swaps)} swap(s) back to original schedule for the new week.")

def reset_weekly_leaves():
    with app.app_context():
        all_leaves = Leave.query.all()
        reverted = 0

        for leave in all_leaves:
            entry = TimetableEntry.query.get(leave.entry_id)
            class_obj = Class.query.get(entry.class_id) if entry else None

            if leave.status == "confirmed" and leave.confirmed_faculty_id and entry and class_obj:
                substitute = Faculty.query.get(leave.confirmed_faculty_id)
                if substitute:
                    create_notification(
                        "faculty", substitute.faculty_id,
                        f"You're no longer covering {class_obj.year} - {class_obj.department} - "
                        f"{class_obj.section}, {entry.day_of_week} Period {entry.period_no} — "
                        f"the week has reset and the schedule is back to normal.",
                        "cover_request"
                    )

            if entry:
                entry.status_color = "normal"

            CoverRequest.query.filter_by(leave_id=leave.leave_id).delete()
            db.session.delete(leave)
            reverted += 1

        db.session.commit()
        print(f"[WEEKLY RESET] Reverted {reverted} leave(s) back to original schedule for the new week.")
        

def create_notification(recipient_type, recipient_id, message, notif_type, reference_id=None):
    notif = UserNotification(
        recipient_type=recipient_type,
        recipient_id=recipient_id,
        message=message,
        notif_type=notif_type,
        reference_id=reference_id
    )
    db.session.add(notif)
    db.session.commit()

def check_pending_leaves():
    with app.app_context():
        pending_leaves = Leave.query.filter(Leave.status != "confirmed").all()
        print(f"[DEBUG] Scheduler ran. Found {len(pending_leaves)} pending leave(s).")

        for leave in pending_leaves:
            already_notified = Notification.query.filter_by(leave_id=leave.leave_id).first()
            if already_notified:
                print(f"[DEBUG] Leave {leave.leave_id} already notified, skipping.")
                continue

            entry = TimetableEntry.query.get(leave.entry_id)
            config = TimetableConfig.query.filter_by(class_id=entry.class_id).first()
            if not config:
                continue

            leave_date_obj = datetime.strptime(leave.leave_date, "%Y-%m-%d")
            start_hour, start_minute = map(int, config.start_time.split(":"))
            minutes_to_add = (entry.period_no - 1) * config.period_duration_minutes
            class_datetime = leave_date_obj.replace(hour=start_hour, minute=start_minute) + timedelta(minutes=minutes_to_add)

            now = datetime.now()
            reminder_time = class_datetime - timedelta(hours=3)

            print(f"[DEBUG] now={now}, reminder_time={reminder_time}, class_datetime={class_datetime}")

            if now >= reminder_time and now < class_datetime:
                faculty = Faculty.query.get(leave.faculty_id)
                class_obj = Class.query.get(entry.class_id)
                class_label = f"{class_obj.year} - {class_obj.department} - {class_obj.section}" if class_obj else "Class"

                print(f"[REMINDER] Faculty {faculty.name if faculty else 'Faculty'}: You haven't assigned anyone for {entry.day_of_week} Period {entry.period_no} at {class_datetime.strftime('%H:%M')}.")
                print(f"[REMINDER TO CC] {faculty.name if faculty else 'Faculty'} is on leave and hasn't assigned a substitute for {entry.day_of_week} Period {entry.period_no}. Class starts at {class_datetime.strftime('%H:%M')}.")

                new_notification = Notification(
                    leave_id=leave.leave_id,
                    sent_to_faculty=True,
                    sent_to_cc=True
                )
                db.session.add(new_notification)

                # Send in-app notification to Faculty on leave
                if faculty:
                    create_notification(
                        "faculty", faculty.faculty_id,
                        f"Reminder: You haven't assigned a substitute for {class_label}, {entry.day_of_week} Period {entry.period_no}.",
                        "leave_reminder", reference_id=leave.leave_id
                    )

                # Send in-app notification to Class Coordinator (CC)
                if class_obj and class_obj.cc_faculty_id:
                    create_notification(
                        "faculty", class_obj.cc_faculty_id,
                        f"Reminder: {faculty.name if faculty else 'Faculty'} is on leave and hasn't assigned a substitute for {class_label}, {entry.day_of_week} Period {entry.period_no}.",
                        "leave_reminder", reference_id=leave.leave_id
                    )

                db.session.commit()
            else:
                print(f"[DEBUG] Not in reminder window yet.")

def get_or_create_faculty_for_admin(admin):
    """Finds this admin's linked Faculty row by email AND college_id, or creates one
    on the fly so the admin can be assigned as CC / course faculty."""
    faculty = Faculty.query.filter_by(college_id=admin.college_id, email=admin.email).first()
    if faculty:
        return faculty
    faculty = Faculty(
        college_id=admin.college_id,
        name=admin.name,
        email=admin.email,
        password_hash=admin.password_hash,
        subject_expertise=admin.department_name or ""
    )
    db.session.add(faculty)
    db.session.commit()
    return faculty


import pandas as pd # or openpyxl to read your uploaded excel file

def run_timetable_solver(excel_file_path):
    # 1. Read data from your 10-sheet Excel file
    # Example: classes_df = pd.read_excel(excel_file_path, sheet_name='Classes')
    # Example: courses_df = pd.read_excel(excel_file_path, sheet_name='Courses')

    # 2. Create the model
    model = cp_model.CpModel()

    # 3. Define decision variables (e.g., assigning courses to time slots and rooms)
    # ... your variable creation logic based on Excel data ...

    # 4. Add constraints (e.g., no room double booking, faculty availability)
    # ... your constraint logic ...

    # 5. Create the solver and solve the model
    solver = cp_model.CpSolver()
    status = solver.solve(model)

    if status == cp_model.OPTIMAL or status == cp_model.FEASIBLE:
        print("Solution found!")
        # Extract schedule from solver and save to Aiven MySQL database
        return True
    else:
        print("No solution found.")
        return False

# ===================================================================
# 2. MULTI-SHEET EXCEL TEMPLATE DOWNLOAD ROUTE
# ===================================================================
@app.route('/download_timetable_template', methods=['GET'])
def download_timetable_template():
    wb = Workbook()
    wb.remove(wb.active)  # Remove standard default sheet

    # Structure mapping: Sheet Name -> Column List
    sheets_structure = {
        "Departments": [
            "Department ID", "Department Name"
        ],
        "Classes": [
            "Class ID", "Academic Year", "Department", "Section", "Semester"
        ],
        "Faculty": [
            "Faculty ID", "Faculty Name", "Department", "Email"
        ],
        "Courses": [
            "Course Code", "Course Name", "Department", "Semester",
            "Course Type", "Hours Per Week", "Periods Per Session"
        ],
        "Faculty Allocation": [
            "Faculty ID", "Course Code", "Class ID", "Hours Per Week"
        ],
        "Rooms": [
            "Room ID", "Room Name", "Room Type", "Capacity"
        ],
        "Time Slots": [
            "Day", "Period", "Start Time", "End Time"
        ],
        "Faculty Availability": [
            "Faculty ID", "Day", "Period", "Availability"
        ],
        "Room Availability": [
            "Room ID", "Day", "Period", "Availability"
        ],
        "Constraints": [
            "Enable", "Constraint Name", "Constraint Type", "Entity Type", "Entity", "Day", "Period", "Value", "Priority"
        ]
    }

    # Header styling constants
    header_font = Font(name="Calibri", size=11, bold=True, color="FFFFFF")
    header_fill = PatternFill(start_color="1F4E79", end_color="1F4E79", fill_type="solid")
    header_alignment = Alignment(horizontal="center", vertical="center", wrap_text=True)
    thin_border = Border(
        left=Side(style='thin', color='D9D9D9'),
        right=Side(style='thin', color='D9D9D9'),
        top=Side(style='thin', color='D9D9D9'),
        bottom=Side(style='medium', color='1F4E79')
    )

    # Build sheets and format headers
    for sheet_name, headers in sheets_structure.items():
        ws = wb.create_sheet(title=sheet_name)
        ws.row_dimensions[1].height = 28
        ws.views.sheetView[0].showGridLines = True

        for col_idx, header in enumerate(headers, 1):
            cell = ws.cell(row=1, column=col_idx, value=header)
            cell.font = header_font
            cell.fill = header_fill
            cell.alignment = header_alignment
            cell.border = thin_border
            
            # Set explicit clean width based on header length
            col_letter = cell.column_letter
            ws.column_dimensions[col_letter].width = max(len(header) + 5, 16)

    # --- Pre-populate Redesigned Constraints Sheet ---
    ws_constraints = wb["Constraints"]
    predefined_constraints = [
        ("FALSE", "Fix Specific Course to Slot", "Fixed Course", "Course", "", "", "", "", "High"),
        ("FALSE", "Fix Specific Class to Slot", "Fixed Class", "Class", "", "", "", "", "High"),
        ("FALSE", "Preferred Course Day", "Preferred Day", "Course", "", "", "", "", "Medium"),
        ("FALSE", "Preferred Course Period", "Preferred Period", "Course", "", "", "", "", "Medium"),
        ("FALSE", "Avoid Course Day", "Avoid Day", "Course", "", "", "", "", "Medium"),
        ("FALSE", "Avoid Course Period", "Avoid Period", "Course", "", "", "", "", "Medium"),
        ("FALSE", "Restrict Multiple Course Sessions per Day", "Course Same Day Restriction", "Course", "", "", "", "", "Medium"),
        ("FALSE", "Maximum Consecutive Teaching Classes", "Maximum Consecutive Classes", "Faculty", "", "", "", "3", "Medium"),
        ("FALSE", "Minimum Gap Between Classes", "Minimum Gap Between Classes", "Faculty", "", "", "", "1", "Medium"),
        ("FALSE", "Spread Course Sessions Across Days", "Spread Course Across Days", "Course", "", "", "", "", "Medium"),
        ("FALSE", "Avoid First Period Schedule", "Avoid First Period", "Faculty", "", "", "", "", "Low"),
        ("FALSE", "Avoid Last Period Schedule", "Avoid Last Period", "Faculty", "", "", "", "", "Low"),
        ("FALSE", "Faculty Preferred Working Day", "Faculty Preferred Day", "Faculty", "", "", "", "", "Medium"),
        ("FALSE", "Faculty Preferred Working Period", "Faculty Preferred Period", "Faculty", "", "", "", "", "Medium"),
        ("FALSE", "Class Preferred Attending Day", "Class Preferred Day", "Class", "", "", "", "", "Medium"),
        ("FALSE", "Class Preferred Attending Period", "Class Preferred Period", "Class", "", "", "", "", "Medium"),
        ("FALSE", "Preferred Lab Session Day", "Preferred Lab Day", "Lab Course", "", "", "", "", "Medium"),
        ("FALSE", "Preferred Lab Session Period", "Preferred Lab Period", "Lab Course", "", "", "", "", "Medium"),
        ("FALSE", "Avoid Lab Session Day", "Avoid Lab Day", "Lab Course", "", "", "", "", "Medium"),
        ("FALSE", "Avoid Lab Session Period", "Avoid Lab Period", "Lab Course", "", "", "", "", "Medium"),
        ("FALSE", "Preferred Laboratory Venue", "Preferred Laboratory/Room", "Lab Course", "", "", "", "", "Medium"),
        ("FALSE", "Avoid Specific Laboratory Venue", "Avoid Specific Laboratory/Room", "Lab Course", "", "", "", "", "Medium"),
    ]

    for r_idx, row_values in enumerate(predefined_constraints, start=2):
        ws_constraints.row_dimensions[r_idx].height = 20
        for c_idx, val in enumerate(row_values, start=1):
            cell = ws_constraints.cell(row=r_idx, column=c_idx, value=val)
            cell.alignment = Alignment(vertical="center")

    # Adjust Constraints column widths
    constraints_widths = [10, 32, 28, 16, 16, 14, 10, 10, 12]
    for col_idx, width in enumerate(constraints_widths, start=1):
        ws_constraints.column_dimensions[ws_constraints.cell(row=1, column=col_idx).column_letter].width = width

    # --- Data Validations (Dropdowns) ---

    # 1. Sheet 4 (Courses) - Course Type Dropdown (Column E)
    ws_courses = wb["Courses"]
    dv_course_type = DataValidation(type="list", formula1='"Theory,Lab"', allow_blank=True)
    ws_courses.add_data_validation(dv_course_type)
    dv_course_type.add("E2:E1000")

    # 2. Sheet 10 (Constraints) Dropdowns
    # Enable Dropdown (Col A)
    dv_enable = DataValidation(type="list", formula1='"TRUE,FALSE"', allow_blank=True)
    ws_constraints.add_data_validation(dv_enable)
    dv_enable.add("A2:A1000")

    # Constraint Type Dropdown (Col C)
    c_type_options = '"Fixed Course,Fixed Class,Preferred Day,Preferred Period,Avoid Day,Avoid Period,Course Same Day Restriction,Maximum Consecutive Classes,Minimum Gap Between Classes,Spread Course Across Days,Avoid First Period,Avoid Last Period,Faculty Preferred Day,Faculty Preferred Period,Class Preferred Day,Class Preferred Period,Preferred Lab Day,Preferred Lab Period,Avoid Lab Day,Avoid Lab Period,Preferred Laboratory/Room,Avoid Specific Laboratory/Room"'
    dv_c_type = DataValidation(type="list", formula1=c_type_options, allow_blank=True)
    ws_constraints.add_data_validation(dv_c_type)
    dv_c_type.add("C2:C1000")

    # Entity Type Dropdown (Col D)
    dv_e_type = DataValidation(type="list", formula1='"Course,Class,Faculty,Lab Course,Room"', allow_blank=True)
    ws_constraints.add_data_validation(dv_e_type)
    dv_e_type.add("D2:D1000")

    # Day Dropdown (Col F)
    dv_day = DataValidation(type="list", formula1='"Monday,Tuesday,Wednesday,Thursday,Friday,Saturday"', allow_blank=True)
    ws_constraints.add_data_validation(dv_day)
    dv_day.add("F2:F1000")

    # Priority Dropdown (Col I)
    dv_priority = DataValidation(type="list", formula1='"High,Medium,Low"', allow_blank=True)
    ws_constraints.add_data_validation(dv_priority)
    dv_priority.add("I2:I1000")

    # Save to memory stream
    output = io.BytesIO()
    wb.save(output)
    output.seek(0)

    return send_file(
        output,
        mimetype="application/vnd.openxmlformats-officedocument.spreadsheetml.sheet",
        as_attachment=True,
        download_name="Timetable_Template.xlsx"
    )

# ===================================================================
# 3. ASYNC TRIGGER & STATUS ENDPOINTS
# ===================================================================

@app.route('/upload_and_generate_timetable_async', methods=['POST'])
def upload_and_generate_timetable_async():
    if 'file' not in request.files:
        return jsonify({"error": "No file uploaded"}), 400

    college_id = request.form.get('college_id', type=int)
    semester_id = request.form.get('semester_id', type=int)
    file = request.files['file']
    file_bytes = file.read()

    job_id = str(uuid.uuid4())
    job = Job(
        job_id=job_id,
        college_id=college_id,
        job_type='timetable_generation',
        status='pending',
        progress_message='Queued for execution'
    )
    db.session.add(job)
    db.session.commit()

    app_ctx = app.app_context()
    thread = threading.Thread(
        target=process_and_solve_job, 
        args=(app_ctx, job_id, college_id, semester_id, file_bytes)
    )
    thread.start()

    return jsonify({"job_id": job_id, "status": "pending"}), 202


@app.route('/job_status/<job_id>', methods=['GET'])
def get_job_status(job_id):
    job = Job.query.get(job_id)
    if not job:
        return jsonify({"error": "Job not found"}), 404

    return jsonify({
        "job_id": job.job_id,
        "status": job.status,
        "progress_message": job.progress_message,
        "result": job.result,
        "error": job.error
    }), 200

def process_and_solve_job(app_ctx, job_id, college_id, semester_id, file_bytes):
    with app_ctx:
        job = Job.query.filter_by(job_id=job_id).first()
        if not job:
            return

        try:
            # 1. Update status to running
            job.status = 'processing'
            job.progress_message = 'Running OR-Tools solver...'
            db.session.commit()

            # 2. Save file bytes temporarily so pandas can read it
            temp_file_path = f"temp_job_{job_id}.xlsx"
            with open(temp_file_path, "wb") as f:
                f.write(file_bytes)

            # --- STRICT ROW COUNT VALIDATION & DB SYNC ---
            excel_file = pd.ExcelFile(temp_file_path)
            required_sheets = ['Faculty', 'Rooms', 'Courses', 'Departments', 'Classes', 'Faculty Allocation', 'Time Slots']
            for sheet in required_sheets:
                if sheet not in excel_file.sheet_names:
                    raise ValueError(f"Missing required sheet: {sheet}")
                
                df = pd.read_excel(excel_file, sheet_name=sheet)
                df_cleaned = df.dropna(how='all')
                if len(df_cleaned) == 0:
                    raise ValueError(f"The sheet '{sheet}' contains no data rows.")

            # Load DataFrames
            sheets_data = {s: pd.read_excel(excel_file, sheet_name=s).dropna(how='all') for s in required_sheets}
            excel_file.close() # Close file handle

            df_departments = sheets_data["Departments"]
            df_classes = sheets_data["Classes"]
            df_faculty = sheets_data["Faculty"]
            df_courses = sheets_data["Courses"]
            df_alloc = sheets_data["Faculty Allocation"]
            df_slots = sheets_data["Time Slots"]

            # 1. Map Departments ID -> Name
            dept_map = {}
            for _, r in df_departments.iterrows():
                d_id = str(r.get("Department ID", "")).strip()
                d_name = str(r.get("Department Name", "")).strip()
                if d_id:
                    dept_map[d_id] = d_name

            # 2. Sync Faculty
            fac_map = {} # Excel Faculty ID -> DB Faculty object
            for _, r in df_faculty.iterrows():
                excel_fac_id = str(r.get("Faculty ID", "")).strip()
                fac_name = str(r.get("Faculty Name", "")).strip()
                fac_email = str(r.get("Email", "")).strip()
                if not excel_fac_id:
                    continue
                fac = None
                if fac_email:
                    fac = Faculty.query.filter_by(college_id=college_id, email=fac_email).first()
                if not fac:
                    # Auto-create account: set temporary password to his email ID
                    fac_pw_hash = bcrypt.generate_password_hash(fac_email).decode('utf-8')
                    fac = Faculty(
                        college_id=college_id,
                        name=fac_name,
                        email=fac_email,
                        password_hash=fac_pw_hash
                    )
                    db.session.add(fac)
                    db.session.flush()
                fac_map[excel_fac_id] = fac

            # 3. Sync Classes
            class_map = {} # Excel Class ID -> DB Class object
            for _, r in df_classes.iterrows():
                excel_class_id = str(r.get("Class ID", "")).strip()
                excel_dept_id = str(r.get("Department", "")).strip()
                dept_name = dept_map.get(excel_dept_id, excel_dept_id)
                year = str(r.get("Academic Year", "")).strip()
                section = str(r.get("Section", "")).strip()
                if not excel_class_id:
                    continue
                cls = Class.query.filter_by(
                    college_id=college_id,
                    semester_id=semester_id,
                    year=year,
                    section=section,
                    department=dept_name
                ).first()
                if not cls:
                    cls = Class(
                        college_id=college_id,
                        semester_id=semester_id,
                        year=year,
                        section=section,
                        department=dept_name
                    )
                    db.session.add(cls)
                    db.session.flush()
                class_map[excel_class_id] = cls

            # 3b. Sync Rooms
            df_rooms = sheets_data.get("Rooms", pd.DataFrame())
            for _, r in df_rooms.iterrows():
                r_name = str(r.get("Room Name", "") or r.get("Room ID", "")).strip()
                r_type = str(r.get("Room Type", "Lecture")).strip()
                cap = r.get("Capacity")
                cap_val = int(cap) if pd.notna(cap) else None
                if r_name:
                    existing_room = Room.query.filter_by(college_id=college_id, room_name=r_name).first()
                    if not existing_room:
                        new_room = Room(college_id=college_id, room_name=r_name, room_type=r_type, capacity=cap_val)
                        db.session.add(new_room)
            db.session.flush()

            # 4. Sync Courses
            course_names = {}
            for _, r in df_courses.iterrows():
                c_code = str(r.get("Course Code", "")).strip()
                c_name = str(r.get("Course Name", "")).strip()
                if c_code:
                    course_names[c_code] = c_name

            course_map = {} # (excel_course_code, DB_class_id) -> DB Course object
            for _, r in df_alloc.iterrows():
                excel_fac_id = str(r.get("Faculty ID", "")).strip()
                excel_course_code = str(r.get("Course Code", "")).strip()
                excel_class_id = str(r.get("Class ID", "")).strip()
                
                db_class = class_map.get(excel_class_id)
                if not db_class:
                    continue
                
                db_fac = fac_map.get(excel_fac_id)
                db_fac_id = db_fac.faculty_id if db_fac else None
                
                course_name = course_names.get(excel_course_code, excel_course_code)
                
                course = Course.query.filter_by(
                    class_id=db_class.class_id,
                    course_name=course_name
                ).first()
                
                if not course:
                    course = Course(
                        class_id=db_class.class_id,
                        course_name=course_name,
                        course_code=excel_course_code,
                        faculty_id=db_fac_id
                    )
                    db.session.add(course)
                    db.session.flush()
                else:
                    if not course.course_code:
                        course.course_code = excel_course_code
                    if db_fac_id and not course.faculty_id:
                        course.faculty_id = db_fac_id
                    db.session.flush()
                
                course_map[(excel_course_code, db_class.class_id)] = course

            db.session.commit()

            # 5. Extract Time Slot times
            slot_times = {} # (day_lower, period_no) -> (start_time, end_time)
            for _, r in df_slots.iterrows():
                day = str(r.get("Day", "")).strip().lower()
                period = int(r.get("Period", 0))
                start_t = r.get("Start Time")
                end_t = r.get("End Time")
                if pd.notna(start_t):
                    if hasattr(start_t, 'strftime'):
                        start_t = start_t.strftime('%H:%M')
                    else:
                        start_t = str(start_t).strip()
                else:
                    start_t = None
                if pd.notna(end_t):
                    if hasattr(end_t, 'strftime'):
                        end_t = end_t.strftime('%H:%M')
                    else:
                        end_t = str(end_t).strip()
                else:
                    end_t = None
                
                if day and period:
                    slot_times[(day, period)] = (start_t, end_t)

            # -----------------------------------

            # 6. Call the OR-Tools solver
            result = generate_timetable_with_ortools(temp_file_path)

            # Clean up temporary file
            try:
                if os.path.exists(temp_file_path):
                    os.remove(temp_file_path)
            except PermissionError:
                pass  # Windows file lock — file will be cleaned up later

            if result["status"] == "success":
                schedule_entries = result["schedule"]
                
                # Clear old entries for relevant synced classes to avoid duplicate timetable entries
                synced_class_ids = [c.class_id for c in class_map.values()]
                if synced_class_ids:
                    delete_timetable_entries_cascade(class_ids=synced_class_ids)
                    db.session.commit()

                # Day name converter for day_of_week MON, TUE, etc.
                day_name_map = {
                    'monday': 'MON',
                    'tuesday': 'TUE',
                    'wednesday': 'WED',
                    'thursday': 'THU',
                    'friday': 'FRI',
                    'saturday': 'SAT',
                    'sunday': 'SUN'
                }

                # Save each generated entry to TimetableEntry
                for entry in schedule_entries:
                    excel_class_id = str(entry['class_id'])
                    excel_course_code = str(entry['course_code'])
                    day_full = str(entry['day']).strip()
                    period_no = int(entry['period'])
                    
                    db_class = class_map.get(excel_class_id)
                    if not db_class:
                        continue
                        
                    db_course = course_map.get((excel_course_code, db_class.class_id))
                    db_course_id = db_course.course_id if db_course else None
                    
                    # Convert day name to short code
                    day_short = day_name_map.get(day_full.lower(), day_full[:3].upper())
                    
                    # Look up start and end times from slot_times
                    start_t, end_t = slot_times.get((day_full.lower(), period_no), (None, None))
                    
                    new_timetable_entry = TimetableEntry(
                        class_id=db_class.class_id,
                        day_of_week=day_short,
                        period_no=period_no,
                        course_id=db_course_id,
                        start_time=start_t,
                        end_time=end_t,
                        entry_type='period',
                        status_color='normal'
                    )
                    db.session.add(new_timetable_entry)
                
                db.session.commit()

                # 6. Mark job as complete
                job.status = 'success'
                job.progress_message = 'Timetable generated successfully!'
                job.result = {
                    "created": {
                        "classes": len(schedule_entries),
                        "courses": len(set(e['course_code'] for e in schedule_entries)),
                        "rooms": len(set(e['room_id'] for e in schedule_entries))
                    }
                }
                db.session.commit()
            else:
                job.status = 'failed'
                job.error = result["error"]
                db.session.commit()

        except Exception as e:
            db.session.rollback()
            # Clean up temp file if it exists in case of failure
            try:
                if 'temp_file_path' in locals() and os.path.exists(temp_file_path):
                    os.remove(temp_file_path)
            except PermissionError:
                pass  # Windows file lock — file will be cleaned up later
                
            try:
                job.status = 'failed'
                job.error = str(e)
                db.session.commit()
            except Exception as commit_err:
                db.session.rollback()
                print(f"Error marking job as failed: {commit_err}")

    #timetable generator\







@app.route("/admin_ensure_faculty_identity", methods=["POST"])
def admin_ensure_faculty_identity():
    data = request.get_json()
    admin = Admin.query.get(data["admin_id"])
    if not admin:
        return jsonify({"error": "Admin not found"}), 404

    faculty = get_or_create_faculty_for_admin(admin)
    return jsonify({
        "faculty_id": faculty.faculty_id,
        "faculty_name": faculty.name
    })

@app.route("/")
def home():
    return "Timetable App Backend is Running!"

@app.route("/admin_invite_cc", methods=["POST"])
def admin_invite_cc():
    data = request.get_json()

    class_obj = Class.query.get(data["class_id"])
    if not class_obj:
        return jsonify({"error": "Class not found"}), 404

    if class_obj.cc_faculty_id is not None:
        return jsonify({"error": "This class already has a CC assigned"}), 400

    existing = CCRequest.query.filter_by(
        class_id=data["class_id"],
        faculty_id=data["faculty_id"],
        status="pending"
    ).first()
    if existing:
        return jsonify({"error": "This faculty already has a pending CC invite for this class"}), 400

    new_request = CCRequest(
        class_id=data["class_id"],
        faculty_id=data["faculty_id"],
        initiated_by="admin",
        status="pending"
    )
    db.session.add(new_request)
    db.session.commit()

    create_notification(
        "faculty", data["faculty_id"],
        f"You've been invited to be CC for {class_obj.year} - {class_obj.department} - {class_obj.section}",
        "cc_invite", reference_id=new_request.cc_request_id
    )

    return jsonify({"message": "Invitation sent to faculty!", "request_id": new_request.cc_request_id})


@app.route("/request_cc", methods=["POST"])
def request_cc():
    data = request.get_json()

    class_obj = Class.query.get(data["class_id"])
    if not class_obj:
        return jsonify({"error": "Class not found"}), 404
    if class_obj.cc_faculty_id is not None:
        return jsonify({"error": "This class already has a CC assigned"}), 400

    existing = CCRequest.query.filter_by(
        class_id=data["class_id"],
        faculty_id=data["faculty_id"],
        status="pending"
    ).first()
    if existing:
        return jsonify({"error": "You already have a pending CC request for this class"}), 400

    new_request = CCRequest(
        class_id=data["class_id"],
        faculty_id=data["faculty_id"],
        initiated_by="faculty",
        status="pending"
    )
    db.session.add(new_request)
    db.session.commit()

    faculty = Faculty.query.get(data["faculty_id"])
    admin = Admin.query.filter_by(college_id=class_obj.college_id).first()
    if admin:
        create_notification(
            "admin", admin.admin_id,
            f"{faculty.name} requested to be CC for {class_obj.year} - {class_obj.department} - {class_obj.section}",
            "cc_invite", reference_id=new_request.cc_request_id
        )

    return jsonify({"message": "CC request sent successfully!", "request_id": new_request.cc_request_id})

@app.route("/admin_invite_course_faculty", methods=["POST"])
def admin_invite_course_faculty():
    data = request.get_json()

    course = Course.query.get(data["course_id"])
    if not course:
        return jsonify({"error": "Course not found"}), 404

    if course.faculty_id is not None:
        return jsonify({"error": "This course already has a faculty assigned"}), 400

    existing = CourseFacultyRequest.query.filter_by(
        course_id=data["course_id"],
        faculty_id=data["faculty_id"],
        status="pending"
    ).first()
    if existing:
        return jsonify({"error": "This faculty already has a pending invite for this course"}), 400

    new_request = CourseFacultyRequest(
        course_id=data["course_id"],
        faculty_id=data["faculty_id"],
        initiated_by="admin",
        status="pending"
    )
    db.session.add(new_request)
    db.session.commit()

    class_obj = Class.query.get(course.class_id)
    class_label = f"{class_obj.year} - {class_obj.department} - {class_obj.section}"
    create_notification(
        "faculty", data["faculty_id"],
        f"You've been invited to teach {course.course_name} for {class_label}",
        "course_invite", reference_id=new_request.cf_request_id
    )

    return jsonify({"message": "Invitation sent to faculty!", "request_id": new_request.cf_request_id})


@app.route("/request_course_faculty", methods=["POST"])
def request_course_faculty():
    data = request.get_json()

    course = Course.query.get(data["course_id"])
    if not course:
        return jsonify({"error": "Course not found"}), 404
    if course.faculty_id is not None:
        return jsonify({"error": "This course already has a faculty assigned"}), 400

    existing = CourseFacultyRequest.query.filter_by(
        course_id=data["course_id"],
        faculty_id=data["faculty_id"],
        status="pending"
    ).first()
    if existing:
        return jsonify({"error": "You already have a pending request for this course"}), 400

    new_request = CourseFacultyRequest(
        course_id=data["course_id"],
        faculty_id=data["faculty_id"],
        initiated_by="faculty",
        status="pending"
    )
    db.session.add(new_request)
    db.session.commit()

    faculty = Faculty.query.get(data["faculty_id"])
    class_obj = Class.query.get(course.class_id)
    class_label = f"{class_obj.year} - {class_obj.department} - {class_obj.section}"
    admin = Admin.query.filter_by(college_id=class_obj.college_id).first()
    if admin:
        create_notification(
            "admin", admin.admin_id,
            f"{faculty.name} requested to teach {course.course_name} for {class_label}",
            "course_invite", reference_id=new_request.cf_request_id
        )

    return jsonify({"message": "Course faculty request sent!", "request_id": new_request.cf_request_id})


@app.route("/resolve_course_faculty_request", methods=["POST"])
def resolve_course_faculty_request():
    data = request.get_json()
    cf_request = CourseFacultyRequest.query.get(data["cf_request_id"])

    if not cf_request:
        return jsonify({"error": "Request not found"}), 404

    course = Course.query.get(cf_request.course_id)
    class_obj = Class.query.get(course.class_id)
    class_label = f"{class_obj.year} - {class_obj.department} - {class_obj.section}"

    if cf_request.status != "pending":
        if cf_request.status == "rejected" and course.faculty_id is not None:
            other = Faculty.query.get(course.faculty_id)
            other_name = other.name if other else "another faculty member"
            return jsonify({
                "error": f"This invitation is no longer available — {other_name} "
                         f"was already assigned to teach {course.course_name} for {class_label}."
            }), 400
        return jsonify({"error": "This invitation is no longer available."}), 400

    if data["decision"] == "accepted" and course.faculty_id is not None:
        return jsonify({"error": "This course already has a faculty assigned"}), 400

    cf_request.status = data["decision"]
    cf_request.resolved_at = db.func.now()

    if data["decision"] == "accepted":
        course.faculty_id = cf_request.faculty_id

        other_pending = CourseFacultyRequest.query.filter(
            CourseFacultyRequest.course_id == cf_request.course_id,
            CourseFacultyRequest.cf_request_id != cf_request.cf_request_id,
            CourseFacultyRequest.status == "pending"
        ).all()
        for other in other_pending:
            other.status = "rejected"
            other.resolved_at = db.func.now()
            create_notification(
                "faculty", other.faculty_id,
                f"Your request to teach {course.course_name} for {class_label} was "
                f"automatically declined because another faculty was assigned first.",
                "course_response"
            )

    db.session.commit()

    faculty = Faculty.query.get(cf_request.faculty_id)
    create_notification(
        "faculty", faculty.faculty_id,
        f"Your request to teach {course.course_name} for {class_label} was {data['decision']}",
        "course_response"
    )
    if cf_request.initiated_by == "faculty":
        admin = Admin.query.filter_by(college_id=class_obj.college_id).first()
        if admin:
            create_notification(
                "admin", admin.admin_id,
                f"{faculty.name}'s request to teach {course.course_name} for {class_label} was {data['decision']}",
                "course_response"
            )
    return jsonify({"message": f"Course faculty request {data['decision']} successfully!"})

@app.route("/resolve_cc_request", methods=["POST"])
def resolve_cc_request():
    data = request.get_json()

    cc_request = CCRequest.query.get(data["cc_request_id"])
    if not cc_request:
        return jsonify({"error": "Request not found"}), 404

    class_obj = Class.query.get(cc_request.class_id)
    class_label = f"{class_obj.year} - {class_obj.department} - {class_obj.section}"

    if cc_request.status != "pending":
        if cc_request.status == "rejected" and class_obj.cc_faculty_id is not None:
            other = Faculty.query.get(class_obj.cc_faculty_id)
            other_name = other.name if other else "another faculty member"
            return jsonify({
                "error": f"This invitation is no longer available — {other_name} "
                         f"was already assigned as CC for {class_label}."
            }), 400
        return jsonify({"error": "This invitation is no longer available."}), 400

    if data["decision"] == "accepted" and class_obj.cc_faculty_id is not None:
        return jsonify({"error": "This class already has a CC assigned"}), 400

    cc_request.status = data["decision"]
    cc_request.resolved_at = db.func.now()

    if data["decision"] == "accepted":
        old_classes = Class.query.filter(
            Class.cc_faculty_id == cc_request.faculty_id,
            Class.class_id != cc_request.class_id
        ).all()
        for old_class in old_classes:
            old_class.cc_faculty_id = None

        class_obj.cc_faculty_id = cc_request.faculty_id

        other_pending = CCRequest.query.filter(
            CCRequest.class_id == cc_request.class_id,
            CCRequest.cc_request_id != cc_request.cc_request_id,
            CCRequest.status == "pending"
        ).all()
        for other in other_pending:
            other.status = "rejected"
            other.resolved_at = db.func.now()
            create_notification(
                "faculty", other.faculty_id,
                f"Your CC request for {class_label} was automatically declined "
                f"because another faculty was assigned first.",
                "cc_response"
            )

    db.session.commit()

    faculty = Faculty.query.get(cc_request.faculty_id)
    create_notification(
        "faculty", faculty.faculty_id,
        f"Your CC request for {class_label} was {data['decision']}",
        "cc_response"
    )
    if cc_request.initiated_by == "faculty":
        admin = Admin.query.filter_by(college_id=class_obj.college_id).first()
        if admin:
            create_notification(
                "admin", admin.admin_id,
                f"{faculty.name}'s CC request for {class_label} was {data['decision']}",
                "cc_response"
            )

    return jsonify({"message": f"CC request {data['decision']} successfully!"})


@app.route("/faculty_cc_invites/<int:faculty_id>", methods=["GET"])
def get_faculty_cc_invites(faculty_id):
    requests = CCRequest.query.filter_by(faculty_id=faculty_id, status="pending", initiated_by="admin").all()
    result = []
    for r in requests:
        class_obj = Class.query.get(r.class_id)
        result.append({
            "cc_request_id": r.cc_request_id,
            "class_id": r.class_id,
            "class_name": f"{class_obj.year} - {class_obj.section}"
        })
    return jsonify(result)

@app.route("/faculty_related_classes/<int:faculty_id>", methods=["GET"])
def get_faculty_related_classes(faculty_id):
    faculty = Faculty.query.get(faculty_id)
    if not faculty:
        return jsonify([])

    active_semester = Semester.query.filter_by(college_id=faculty.college_id, is_active=True).first()
    semester_id = active_semester.semester_id if active_semester else None

    cc_classes = Class.query.filter_by(cc_faculty_id=faculty_id, semester_id=semester_id).all()

    course_class_ids = db.session.query(Course.class_id).filter_by(faculty_id=faculty_id).distinct().all()
    course_class_ids = [c[0] for c in course_class_ids]
    course_classes = Class.query.filter(
        Class.class_id.in_(course_class_ids),
        Class.semester_id == semester_id
    ).all() if course_class_ids else []

    merged = {c.class_id: c for c in cc_classes + course_classes}.values()

    result = []
    for c in merged:
        result.append({
            "class_id": c.class_id,
            "year": c.year,
            "section": c.section,
            "department": c.department,
            "cc_faculty_id": c.cc_faculty_id
        })
    return jsonify(result)

@app.route("/add_college", methods=["POST"])
def add_college():
    data = request.get_json()
    new_college = College(
        college_name=data["college_name"],
        college_code=data["college_code"]
    )
    db.session.add(new_college)
    db.session.commit()
    return jsonify({"message": "College added successfully!"})

@app.route("/colleges", methods=["GET"])
def get_colleges():
    all_colleges = College.query.all()
    result = []
    for c in all_colleges:
        result.append({
            "college_id": c.college_id,
            "college_name": c.college_name,
            "college_code": c.college_code
        })
    return jsonify(result)

@app.route("/register_admin", methods=["POST"])
def register_admin():
    data = request.get_json()

    college = College.query.filter_by(college_code=data["college_code"].strip()).first()
    if not college:
        return jsonify({"error": "Invalid college code"}), 400

    existing_admin = Admin.query.filter_by(college_id=college.college_id, email=data["email"].strip()).first()
    if existing_admin:
        return jsonify({"error": "Email already registered for this college code"}), 400

    hashed_pw = bcrypt.generate_password_hash(data["password"]).decode('utf-8')

    new_admin = Admin(
        college_id=college.college_id,
        name=data["name"].strip(),
        email=data["email"].strip(),
        password_hash=hashed_pw,
        department_name=data.get("department_name", "").strip()
    )
    db.session.add(new_admin)
    db.session.commit()

    return jsonify({"message": "Admin registered successfully!"})

@app.route("/login_admin", methods=["POST"])
def login_admin():
    data = request.get_json()
    college_code = data.get("college_code", "").strip()
    email = data.get("email", "").strip()
    password = data.get("password", "")

    if not college_code:
        return jsonify({"error": "College code is required for login."}), 400

    college = College.query.filter_by(college_code=college_code).first()
    if not college:
        return jsonify({"error": f"College code '{college_code}' not found."}), 404

    admin = Admin.query.filter_by(college_id=college.college_id, email=email).first()
    if not admin:
        return jsonify({"error": f"Admin account for '{email}' not found under college code '{college_code}'."}), 404

    if bcrypt.check_password_hash(admin.password_hash, password):
        return jsonify({
            "message": "Login successful",
            "admin_id": admin.admin_id,
            "name": admin.name,
            "college_id": admin.college_id
        })
    else:
        return jsonify({"error": "Incorrect password"}), 401

@app.route("/register_faculty", methods=["POST"])
def register_faculty():
    data = request.get_json()

    college = College.query.filter_by(college_code=data["college_code"].strip()).first()
    if not college:
        return jsonify({"error": "Invalid college code"}), 400

    existing_faculty = Faculty.query.filter_by(college_id=college.college_id, email=data["email"].strip()).first()
    if existing_faculty:
        return jsonify({"error": "Email already registered for this college code"}), 400

    hashed_pw = bcrypt.generate_password_hash(data["password"]).decode('utf-8')

    new_faculty = Faculty(
        college_id=college.college_id,
        name=data["name"].strip(),
        email=data["email"].strip(),
        password_hash=hashed_pw,
        subject_expertise=data.get("subject_expertise", "").strip()
    )
    db.session.add(new_faculty)
    db.session.commit()

    return jsonify({"message": "Faculty registered successfully!"})

@app.route("/login_faculty", methods=["POST"])
def login_faculty():
    data = request.get_json()
    college_code = data.get("college_code", "").strip()
    email = data.get("email", "").strip()
    password = data.get("password", "")

    if not college_code:
        return jsonify({"error": "College code is required for login."}), 400

    college = College.query.filter_by(college_code=college_code).first()
    if not college:
        return jsonify({"error": f"College code '{college_code}' not found."}), 404

    faculty = Faculty.query.filter_by(college_id=college.college_id, email=email).first()
    if not faculty:
        return jsonify({"error": f"Faculty account for '{email}' not found under college code '{college_code}'."}), 404

    if bcrypt.check_password_hash(faculty.password_hash, password):
        return jsonify({
            "message": "Login successful",
            "faculty_id": faculty.faculty_id,
            "name": faculty.name,
            "college_id": faculty.college_id
        })
    else:
        return jsonify({"error": "Incorrect password"}), 401

@app.route("/classes/<int:college_id>", methods=["GET"])
def get_classes(college_id):
    active_semester = Semester.query.filter_by(college_id=college_id, is_active=True).first()

    if active_semester:
        classes = Class.query.filter_by(college_id=college_id, semester_id=active_semester.semester_id).all()
    else:
        # No active semester set yet - show classes with no semester (legacy/fallback)
        classes = Class.query.filter_by(college_id=college_id, semester_id=None).all()

    result = []
    for c in classes:
        result.append({
            "class_id": c.class_id,
            "year": c.year,
            "section": c.section,
            "department": c.department,
            "cc_faculty_id": c.cc_faculty_id
        })
    return jsonify(result)

@app.route("/cc_requests/<int:class_id>", methods=["GET"])
def get_cc_requests(class_id):
    requests = CCRequest.query.filter_by(class_id=class_id, status="pending").all()
    result = []
    for r in requests:
        faculty = Faculty.query.get(r.faculty_id)
        result.append({
            "cc_request_id": r.cc_request_id,
            "faculty_id": r.faculty_id,
            "faculty_name": faculty.name,
            "status": r.status
        })
    return jsonify(result)
    
@app.route("/add_course", methods=["POST"])
def add_course():
    data = request.get_json()

    existing = Course.query.filter_by(
        class_id=data["class_id"],
        course_name=data["course_name"].strip()
    ).first()
    if existing:
        return jsonify({"message": "Course already exists", "course_id": existing.course_id})

    new_course = Course(
        class_id=data["class_id"],
        course_name=data["course_name"].strip()
    )
    db.session.add(new_course)
    db.session.commit()
    return jsonify({"message": "Course added successfully!", "course_id": new_course.course_id})

@app.route("/set_timetable_config", methods=["POST"])
def set_timetable_config():
    data = request.get_json()
    new_config = TimetableConfig(
        class_id=data["class_id"],
        periods_per_day=data["periods_per_day"],
        period_duration_minutes=data["period_duration_minutes"],
        start_time=data["start_time"]
    )
    db.session.add(new_config)
    db.session.commit()
    return jsonify({"message": "Timetable config set successfully!"})

@app.route("/add_timetable_entry", methods=["POST"])
def add_timetable_entry():
    data = request.get_json()
    new_entry = TimetableEntry(
        class_id=data["class_id"],
        day_of_week=data["day_of_week"],
        period_no=data["period_no"],
        course_id=data.get("course_id")
    )
    db.session.add(new_entry)
    db.session.commit()
    return jsonify({"message": "Timetable entry added!", "entry_id": new_entry.entry_id})

@app.route("/timetable/<int:class_id>", methods=["GET"])
def get_timetable(class_id):
    entries = TimetableEntry.query.filter_by(class_id=class_id).all()
    result = []
    for e in entries:
        course_name = None
        faculty_name = None
        faculty_id_result = None

        if e.course_id:
            course = Course.query.get(e.course_id)
            course_name = course.course_name
            if course.faculty_id:
                faculty = Faculty.query.get(course.faculty_id)
                faculty_name = faculty.name
                faculty_id_result = faculty.faculty_id

        if course_name:
            confirmed_leave = Leave.query.filter_by(entry_id=e.entry_id, status="confirmed").first()
            if confirmed_leave and confirmed_leave.confirmed_faculty_id:
                substitute = Faculty.query.get(confirmed_leave.confirmed_faculty_id)
                faculty_name = substitute.name
                faculty_id_result = substitute.faculty_id

        # If this slot was swapped, find the accepted swap it belongs to so
        # the frontend can color-match both halves of the same swap.
        swap_id_result = None
        if e.status_color == "swapped":
            swap = SwapRequest.query.filter(
                db.or_(
                    SwapRequest.requester_entry_id == e.entry_id,
                    SwapRequest.target_entry_id == e.entry_id
                ),
                SwapRequest.status == "accepted"
            ).order_by(SwapRequest.resolved_at.desc()).first()
            if swap:
                swap_id_result = swap.swap_id

        result.append({
            "entry_id": e.entry_id,
            "day_of_week": e.day_of_week,
            "period_no": e.period_no,
            "entry_type": e.entry_type,
            "label": e.label,
            "start_time": e.start_time,
            "end_time": e.end_time,
            "course_id": e.course_id,
            "course_name": course_name,
            "faculty_name": faculty_name,
            "faculty_id": faculty_id_result,
            "status_color": e.status_color,
            "swap_id": swap_id_result
        })

    return jsonify(result)

@app.route("/faculty_timetable/<int:faculty_id>", methods=["GET"])
def get_faculty_timetable(faculty_id):
    try:
        courses = Course.query.filter_by(faculty_id=faculty_id).all()
        course_map = {c.course_id: c for c in courses}
        course_ids = list(course_map.keys())

        substitute_leaves = Leave.query.filter_by(confirmed_faculty_id=faculty_id, status="confirmed").all()
        substitute_entry_ids = [l.entry_id for l in substitute_leaves]

        if not course_ids and not substitute_entry_ids:
            return jsonify([])

        conditions = []
        if course_ids:
            conditions.append(TimetableEntry.course_id.in_(course_ids))
        if substitute_entry_ids:
            conditions.append(TimetableEntry.entry_id.in_(substitute_entry_ids))

        entries = TimetableEntry.query.filter(db.or_(*conditions)).all()

        class_ids = list({e.class_id for e in entries if e.class_id})
        classes = Class.query.filter(Class.class_id.in_(class_ids)).all() if class_ids else []
        class_map = {c.class_id: c for c in classes}

        result = []
        for e in entries:
            course = course_map.get(e.course_id)
            if not course:
                course = Course.query.get(e.course_id) if e.course_id else None
            course_name = course.course_name if course else "Class Period"

            cls = class_map.get(e.class_id)
            class_name = f"{cls.year} - {cls.department} - {cls.section}" if cls else "Class"

            swap_id_result = None
            if e.status_color == "swapped":
                swap = SwapRequest.query.filter(
                    db.or_(
                        SwapRequest.requester_entry_id == e.entry_id,
                        SwapRequest.target_entry_id == e.entry_id
                    ),
                    SwapRequest.status == "accepted"
                ).order_by(SwapRequest.resolved_at.desc()).first()
                if swap:
                    swap_id_result = swap.swap_id

            result.append({
                "entry_id": e.entry_id,
                "class_id": e.class_id,
                "class_name": class_name,
                "department": cls.department if cls else "",
                "section": cls.section if cls else "",
                "year": cls.year if cls else "",
                "room_number": cls.room_number if cls else "Not Set",
                "day_of_week": e.day_of_week,
                "period_no": e.period_no,
                "entry_type": e.entry_type,
                "label": e.label,
                "start_time": e.start_time,
                "end_time": e.end_time,
                "course_id": e.course_id,
                "course_name": course_name,
                "faculty_name": None,
                "faculty_id": faculty_id,
                "status_color": e.status_color,
                "swap_id": swap_id_result
            })

        return jsonify(result)
    except Exception as err:
        print(f"Error in get_faculty_timetable: {err}")
        return jsonify([])

@app.route("/verify_college_code/<string:code>", methods=["GET"])
def verify_college_code(code):
    college = College.query.filter_by(college_code=code).first()
    if not college:
        return jsonify({"error": "Invalid college code"}), 404

    return jsonify({
        "college_id": college.college_id,
        "college_name": college.college_name
    })

@app.route("/departments/<int:college_id>", methods=["GET"])
def get_departments(college_id):
    active_semester = Semester.query.filter_by(college_id=college_id, is_active=True).first()
    if active_semester:
        classes = Class.query.filter_by(college_id=college_id, semester_id=active_semester.semester_id).all()
    else:
        classes = Class.query.filter_by(college_id=college_id, semester_id=None).all()
    departments = list(set([c.department for c in classes]))
    return jsonify(departments)

@app.route("/classes_by_department/<int:college_id>/<string:department>", methods=["GET"])
def get_classes_by_department(college_id, department):
    active_semester = Semester.query.filter_by(college_id=college_id, is_active=True).first()
    if active_semester:
        classes = Class.query.filter_by(
            college_id=college_id, department=department, semester_id=active_semester.semester_id
        ).all()
    else:
        classes = Class.query.filter_by(
            college_id=college_id, department=department, semester_id=None
        ).all()
    result = []
    for c in classes:
        result.append({
            "class_id": c.class_id,
            "year": c.year,
            "section": c.section
        })
    return jsonify(result)

@app.route("/faculty_related_classes_all/<int:faculty_id>", methods=["GET"])
def get_faculty_related_classes_all(faculty_id):
    cc_classes = Class.query.filter_by(cc_faculty_id=faculty_id).all()

    course_class_ids = db.session.query(Course.class_id).filter_by(faculty_id=faculty_id).distinct().all()
    course_class_ids = [c[0] for c in course_class_ids]
    course_classes = Class.query.filter(Class.class_id.in_(course_class_ids)).all() if course_class_ids else []

    merged = {c.class_id: c for c in cc_classes + course_classes}.values()

    result = []
    for c in merged:
        result.append({
            "class_id": c.class_id,
            "year": c.year,
            "section": c.section,
            "department": c.department,
            "cc_faculty_id": c.cc_faculty_id
        })
    return jsonify(result)

@app.route("/mark_leave", methods=["POST"])
def mark_leave():
    data = request.get_json()

    new_leave = Leave(
        faculty_id=data["faculty_id"],
        entry_id=data["entry_id"],
        leave_date=data["leave_date"]
    )
    db.session.add(new_leave)
    db.session.commit()

    # Highlight the slot as "open" in the timetable
    entry = TimetableEntry.query.get(data["entry_id"])
    entry.status_color = "open_leave"
    db.session.commit()

    faculty = Faculty.query.get(data["faculty_id"])
    if faculty and entry:
        create_notification(
            "class", entry.class_id,
            f"Timetable change: {faculty.name} is on leave for {entry.day_of_week} Period {entry.period_no}.",
            "leave_notice"
        )

    return jsonify({"message": "Leave marked and slot highlighted!", "leave_id": new_leave.leave_id})

@app.route("/cc_mark_leave", methods=["POST"])
def cc_mark_leave():
    data = request.get_json()
    cc_faculty_id = data.get("cc_faculty_id")
    entry_id = data.get("entry_id")
    leave_date = data.get("leave_date")

    if not cc_faculty_id or not entry_id or not leave_date:
        return jsonify({"error": "Missing cc_faculty_id, entry_id, or leave_date"}), 400

    entry = TimetableEntry.query.get(entry_id)
    if not entry:
        return jsonify({"error": "Timetable entry not found"}), 404

    class_obj = Class.query.get(entry.class_id)
    if not class_obj:
        return jsonify({"error": "Class not found"}), 404

    # VERIFICATION: Only the designated Class Coordinator of this class can mark absent
    if class_obj.cc_faculty_id != int(cc_faculty_id):
        return jsonify({"error": "Unauthorized: You are not the Class Coordinator of this class."}), 403

    target_faculty_id = entry.faculty_id
    if not target_faculty_id:
        return jsonify({"error": "No faculty assigned to this slot"}), 400

    # Check if leave is already marked for this entry & date
    existing_leave = Leave.query.filter_by(
        entry_id=entry_id,
        leave_date=leave_date
    ).first()

    if existing_leave:
        return jsonify({"error": "Faculty is already marked absent for this slot."}), 400

    new_leave = Leave(
        faculty_id=target_faculty_id,
        entry_id=entry_id,
        leave_date=leave_date
    )
    db.session.add(new_leave)
    entry.status_color = "open_leave"
    db.session.commit()

    # Notify absent faculty member
    target_faculty = Faculty.query.get(target_faculty_id)
    cc_faculty = Faculty.query.get(cc_faculty_id)
    cc_name = cc_faculty.name if cc_faculty else "Class Coordinator"
    
    if target_faculty:
        create_notification(
            "faculty", target_faculty.faculty_id,
            f"Your Class Coordinator ({cc_name}) marked you absent for {entry.day_of_week} Period {entry.period_no} in {class_obj.year} - {class_obj.department} Sec {class_obj.section} on {leave_date}.",
            "leave_notice"
        )
        create_notification(
            "class", class_obj.class_id,
            f"Timetable change: {target_faculty.name if target_faculty else 'Faculty'} is on leave for {entry.day_of_week} Period {entry.period_no}.",
            "leave_notice"
        )

    return jsonify({
        "message": f"Faculty {target_faculty.name if target_faculty else ''} marked absent for this slot!",
        "leave_id": new_leave.leave_id
    })

@app.route("/unmark_leave", methods=["POST"])
def unmark_leave():
    data = request.get_json()
    leave = Leave.query.filter_by(
        entry_id=data["entry_id"],
        faculty_id=data["faculty_id"]
    ).first()

    if not leave:
        return jsonify({"error": "No leave found to unmark"}), 404

    entry = TimetableEntry.query.get(leave.entry_id)
    class_obj = Class.query.get(entry.class_id)

    # If a substitute was already confirmed, let them know they're no
    # longer covering this period before we clear it.
    if leave.status == "confirmed" and leave.confirmed_faculty_id:
        substitute = Faculty.query.get(leave.confirmed_faculty_id)
        if substitute:
            create_notification(
                "faculty", substitute.faculty_id,
                f"You're no longer covering {class_obj.year} - {class_obj.department} - "
                f"{class_obj.section}, {entry.day_of_week} Period {entry.period_no} — "
                f"the original leave was cancelled.",
                "cover_request"
            )

    # Notify anyone who had a pending (not-yet-confirmed) volunteer offer too
    pending_covers = CoverRequest.query.filter_by(
        leave_id=leave.leave_id, status="pending"
    ).all()
    for cr in pending_covers:
        volunteer = Faculty.query.get(cr.requesting_faculty_id)
        create_notification(
            "faculty", volunteer.faculty_id,
            f"The {class_obj.year} - {class_obj.department} - {class_obj.section}, "
            f"{entry.day_of_week} Period {entry.period_no} slot you volunteered "
            f"for is no longer open — the leave was cancelled.",
            "cover_request"
        )

    # Restore the slot to how it was before the leave — normal color,
    # original faculty's period, no substitute.
    entry.status_color = "normal"

    # Delete ALL cover requests tied to this leave (pending, accepted, or
    # rejected) — otherwise the foreign key to leave_id blocks deleting
    # the leave row itself.
    CoverRequest.query.filter_by(leave_id=leave.leave_id).delete()

    db.session.delete(leave)
    db.session.commit()

    return jsonify({"message": "Leave unmarked — the period is back to normal."})

@app.route("/send_cover_request", methods=["POST"])
def send_cover_request():
    data = request.get_json()

    new_cover_req = CoverRequest(
        leave_id=data["leave_id"],
        requesting_faculty_id=data["requesting_faculty_id"]
    )
    db.session.add(new_cover_req)
    db.session.commit()

    # Update leave status to show requests are coming in
    leave = Leave.query.get(data["leave_id"])
    leave.status = "pending_requests"
    db.session.commit()

    entry = TimetableEntry.query.get(leave.entry_id)
    class_obj = Class.query.get(entry.class_id)
    volunteer = Faculty.query.get(data["requesting_faculty_id"])
    course = Course.query.get(entry.course_id) if entry.course_id else None

    create_notification(
        "faculty", leave.faculty_id,
        f"{volunteer.name} volunteered to cover your "
        f"{class_obj.year} - {class_obj.department} - {class_obj.section}, "
        f"{entry.day_of_week} Period {entry.period_no}"
        f"{f' ({course.course_name})' if course else ''}",
        "cover_request", reference_id=new_cover_req.cover_req_id
    )

    return jsonify({"message": "Cover request sent!", "cover_req_id": new_cover_req.cover_req_id})

@app.route("/cover_requests/<int:leave_id>", methods=["GET"])
def get_cover_requests(leave_id):
    requests = CoverRequest.query.filter_by(leave_id=leave_id, status="pending").all()
    result = []
    for r in requests:
        faculty = Faculty.query.get(r.requesting_faculty_id)
        result.append({
            "cover_req_id": r.cover_req_id,
            "requesting_faculty_id": r.requesting_faculty_id,
            "faculty_name": faculty.name
        })
    return jsonify(result)

@app.route("/confirm_cover_request", methods=["POST"])
def confirm_cover_request():
    data = request.get_json()

    leave = Leave.query.get(data["leave_id"])
    if not leave:
        return jsonify({"error": "Leave not found"}), 404
    if leave.status == "confirmed":
        return jsonify({"error": "This leave slot is already confirmed"}), 400

    chosen_request = CoverRequest.query.get(data["cover_req_id"])
    chosen_request.status = "accepted"

    other_requests = CoverRequest.query.filter(
        CoverRequest.leave_id == data["leave_id"],
        CoverRequest.cover_req_id != data["cover_req_id"]
    ).all()
    for r in other_requests:
        r.status = "rejected"

    leave.status = "confirmed"
    leave.confirmed_faculty_id = chosen_request.requesting_faculty_id
    leave.confirmed_by_role = data["confirmed_by_role"]

    entry = TimetableEntry.query.get(leave.entry_id)
    entry.status_color = "confirmed_cover"

    db.session.commit()

    substitute = Faculty.query.get(chosen_request.requesting_faculty_id)
    original_faculty = Faculty.query.get(leave.faculty_id)
    create_notification(
        "faculty", substitute.faculty_id,
        f"You're confirmed to cover {entry.day_of_week} Period {entry.period_no}",
        "cover_confirmed"
    )
    create_notification(
        "faculty", original_faculty.faculty_id,
        f"{substitute.name} will cover your {entry.day_of_week} Period {entry.period_no}",
        "cover_confirmed"
    )
    create_notification(
        "class", entry.class_id,
        f"Timetable change: {substitute.name} will substitute for {original_faculty.name} on {entry.day_of_week} Period {entry.period_no}",
        "cover_confirmed"
    )

    return jsonify({"message": "Cover request confirmed successfully!"})

@app.route("/send_swap_request", methods=["POST"])
def send_swap_request():
    data = request.get_json()

    new_swap = SwapRequest(
        requester_faculty_id=data["requester_faculty_id"],
        requester_entry_id=data["requester_entry_id"],
        target_faculty_id=data["target_faculty_id"],
        target_entry_id=data["target_entry_id"]
    )
    db.session.add(new_swap)
    db.session.commit()

    requester = Faculty.query.get(data["requester_faculty_id"])
    requester_entry = TimetableEntry.query.get(data["requester_entry_id"])
    target_entry = TimetableEntry.query.get(data["target_entry_id"])
    requester_class = Class.query.get(requester_entry.class_id)
    target_class = Class.query.get(target_entry.class_id)

    requester_slot = (
        f"{requester_class.year} - {requester_class.department} - {requester_class.section}, "
        f"{requester_entry.day_of_week} Period {requester_entry.period_no}"
    )
    target_slot = (
        f"{target_class.year} - {target_class.department} - {target_class.section}, "
        f"{target_entry.day_of_week} Period {target_entry.period_no}"
    )

    create_notification(
        "faculty", data["target_faculty_id"],
        f"{requester.name} wants to swap: their {requester_slot} for your {target_slot}",
        "swap_request", reference_id=new_swap.swap_id
    )

    return jsonify({"message": "Swap request sent!", "swap_id": new_swap.swap_id})

@app.route("/swap_requests/<int:target_faculty_id>", methods=["GET"])
def get_swap_requests(target_faculty_id):
    requests = SwapRequest.query.filter_by(target_faculty_id=target_faculty_id, status="pending").all()
    result = []
    for r in requests:
        requester = Faculty.query.get(r.requester_faculty_id)
        requester_entry = TimetableEntry.query.get(r.requester_entry_id)
        target_entry = TimetableEntry.query.get(r.target_entry_id)
        requester_class = Class.query.get(requester_entry.class_id)
        target_class = Class.query.get(target_entry.class_id)
        requester_course = Course.query.get(requester_entry.course_id) if requester_entry.course_id else None
        target_course = Course.query.get(target_entry.course_id) if target_entry.course_id else None

        result.append({
            "swap_id": r.swap_id,
            "requester_name": requester.name,
            "requester_class": f"{requester_class.year} - {requester_class.department} - {requester_class.section}",
            "requester_slot": f"{requester_entry.day_of_week} Period {requester_entry.period_no}",
            "requester_course": requester_course.course_name if requester_course else None,
            "target_class": f"{target_class.year} - {target_class.department} - {target_class.section}",
            "target_slot": f"{target_entry.day_of_week} Period {target_entry.period_no}",
            "target_course": target_course.course_name if target_course else None,
            "requested_at": r.requested_at.strftime("%Y-%m-%d %H:%M") if r.requested_at else ""
        })
    return jsonify(result)


@app.route("/resolve_swap_request", methods=["POST"])
def resolve_swap_request():
    data = request.get_json()

    swap = SwapRequest.query.get(data["swap_id"])
    if not swap:
        return jsonify({"error": "Swap request not found"}), 404
    if swap.status != "pending":
        return jsonify({"error": "This request has already been resolved"}), 400

    swap.status = data["decision"]
    swap.resolved_at = db.func.now()
    swap.rejection_reason = data.get("rejection_reason", "")

    requester_entry = TimetableEntry.query.get(swap.requester_entry_id)
    target_entry = TimetableEntry.query.get(swap.target_entry_id)
    requester_class = Class.query.get(requester_entry.class_id)
    target_class = Class.query.get(target_entry.class_id)

    if data["decision"] == "accepted":
        requester_entry.course_id, target_entry.course_id = target_entry.course_id, requester_entry.course_id
        requester_entry.status_color = "swapped"
        target_entry.status_color = "swapped"
        
        target = Faculty.query.get(swap.target_faculty_id)
        requester = Faculty.query.get(swap.requester_faculty_id)

        # Notify affected classes
        create_notification(
            "class", requester_entry.class_id,
            f"Timetable change: Your {requester_entry.day_of_week} Period {requester_entry.period_no} class is now handled by {target.name} due to a faculty swap.",
            "swap_update"
        )
        create_notification(
            "class", target_entry.class_id,
            f"Timetable change: Your {target_entry.day_of_week} Period {target_entry.period_no} class is now handled by {requester.name} due to a faculty swap.",
            "swap_update"
        )

    db.session.commit()

    requester = Faculty.query.get(swap.requester_faculty_id)
    requester_slot = (
        f"{requester_class.year} - {requester_class.department} - {requester_class.section}, "
        f"{requester_entry.day_of_week} Period {requester_entry.period_no}"
    )
    target_slot = (
        f"{target_class.year} - {target_class.department} - {target_class.section}, "
        f"{target_entry.day_of_week} Period {target_entry.period_no}"
    )
    create_notification(
        "faculty", requester.faculty_id,
        f"Your swap request ({requester_slot} \u2194 {target_slot}) was {data['decision']}",
        "swap_response"
    )

    return jsonify({"message": f"Swap request {data['decision']} successfully!"})

@app.route("/mark_day_leave", methods=["POST"])
def mark_day_leave():
    data = request.get_json()
    faculty_id = data["faculty_id"]
    leave_date = data["leave_date"]
    day_of_week = data["day_of_week"]  # e.g. "MON"

    # Find all timetable entries where this faculty teaches, on that day
    entries = db.session.query(TimetableEntry).join(Course).filter(
        Course.faculty_id == faculty_id,
        TimetableEntry.day_of_week == day_of_week
    ).all()

    if not entries:
        return jsonify({"message": "No classes found for this faculty on that day"}), 200

    created_leaves = []
    for entry in entries:
        new_leave = Leave(
            faculty_id=faculty_id,
            entry_id=entry.entry_id,
            leave_date=leave_date
        )
        db.session.add(new_leave)
        entry.status_color = "open_leave"
        created_leaves.append(entry.entry_id)

    db.session.commit()

    faculty = Faculty.query.get(faculty_id)
    if faculty:
        for entry in entries:
            create_notification(
                "class", entry.class_id,
                f"Timetable change: {faculty.name} is on leave for {entry.day_of_week} Period {entry.period_no}.",
                "leave_notice"
            )

    return jsonify({
        "message": f"Leave marked for {len(created_leaves)} period(s)",
        "affected_entries": created_leaves
    })

@app.route("/create_semester", methods=["POST"])
def create_semester():
    data = request.get_json()

    # Deactivate all other semesters for this college (only one active at a time)
    Semester.query.filter_by(college_id=data["college_id"]).update({"is_active": False})

    new_semester = Semester(
        college_id=data["college_id"],
        semester_name=data["semester_name"],
        is_active=True
    )
    db.session.add(new_semester)
    db.session.commit()

    return jsonify({"message": "New semester created and activated!", "semester_id": new_semester.semester_id})

@app.route("/semesters/<int:college_id>", methods=["GET"])
def get_semesters(college_id):
    semesters = Semester.query.filter_by(college_id=college_id, is_deleted=False).all()
    result = []
    for s in semesters:
        result.append({
            "semester_id": s.semester_id,
            "semester_name": s.semester_name,
            "is_active": s.is_active
        })
    return jsonify(result)

@app.route("/switch_semester", methods=["POST"])
def switch_semester():
    data = request.get_json()

    Semester.query.filter_by(college_id=data["college_id"]).update({"is_active": False})

    semester = Semester.query.get(data["semester_id"])
    semester.is_active = True
    db.session.commit()

    return jsonify({"message": f"Switched to {semester.semester_name}"})

@app.route("/delete_semester", methods=["POST"])
def delete_semester():
    data = request.get_json()
    semester_id = data.get("semester_id")

    semester = Semester.query.get(semester_id)
    if not semester:
        return jsonify({"error": "Semester not found"}), 404

    semester.is_deleted = True

    # Safely clear timetable entries for classes in this deleted semester
    classes = Class.query.filter_by(semester_id=semester_id).all()
    for c in classes:
        delete_timetable_entries_cascade(class_ids=[c.class_id])

    db.session.commit()
    return jsonify({"message": f"{semester.semester_name} deleted successfully!"})

@app.route("/add_class", methods=["POST"])
def add_class():
    data = request.get_json()

    admin = Admin.query.filter_by(admin_id=data["admin_id"]).first()
    if not admin:
        return jsonify({"error": "Invalid admin"}), 400

    active_semester = Semester.query.filter_by(college_id=admin.college_id, is_active=True).first()

    if not active_semester:
        return jsonify({"error": "Please create a semester first before adding classes"}), 400

    existing = Class.query.filter_by(
        college_id=admin.college_id,
        semester_id=active_semester.semester_id,
        year=data["year"].strip(),
        section=data["section"].strip(),
        department=data["department"].strip()
    ).first()
    if existing:
        return jsonify({"error": "This class already exists in the current semester"}), 400

    new_class = Class(
        college_id=admin.college_id,
        semester_id=active_semester.semester_id,
        year=data["year"].strip(),
        section=data["section"].strip(),
        department=data["department"].strip(),
        room_number=data.get("room_number", "").strip() or None
    )
    db.session.add(new_class)
    db.session.commit()

    return jsonify({"message": "Class created successfully!", "class_id": new_class.class_id})

@app.route("/delete_class", methods=["POST"])
def delete_class():
    data = request.get_json()
    class_id = data.get("class_id")
    class_obj = Class.query.get(class_id)
    if not class_obj:
        return jsonify({"error": "Class not found"}), 404

    # 1. Cascade delete all timetable entries & dependent leaves/swaps/covers
    delete_timetable_entries_cascade(class_ids=[class_id])

    # 2. Delete CourseFacultyRequests tied to courses in this class
    course_ids = [c.course_id for c in Course.query.filter_by(class_id=class_id).all()]
    if course_ids:
        CourseFacultyRequest.query.filter(CourseFacultyRequest.course_id.in_(course_ids)).delete(synchronize_session=False)

    # 3. Delete Course records in this class
    Course.query.filter_by(class_id=class_id).delete(synchronize_session=False)

    # 4. Delete CCRequests for this class
    CCRequest.query.filter_by(class_id=class_id).delete(synchronize_session=False)

    # 5. Delete Holidays for this class
    Holiday.query.filter_by(class_id=class_id).delete(synchronize_session=False)

    # 6. Delete TimetableConfig for this class
    TimetableConfig.query.filter_by(class_id=class_id).delete(synchronize_session=False)

    # 7. Delete Class record
    db.session.delete(class_obj)
    db.session.commit()
    return jsonify({"message": "Class deleted successfully!"})

@app.route("/delete_course", methods=["POST"])
def delete_course():
    data = request.get_json()
    course_id = data.get("course_id")
    course = Course.query.get(course_id)
    if not course:
        return jsonify({"error": "Course not found"}), 404

    # Clear timetable entries referencing this course
    entries = TimetableEntry.query.filter_by(course_id=course_id).all()
    for e in entries:
        e.course_id = None
        e.entry_type = 'free'

    # Delete requests tied to this course
    CourseFacultyRequest.query.filter_by(course_id=course_id).delete(synchronize_session=False)

    db.session.delete(course)
    db.session.commit()
    return jsonify({"message": "Course deleted successfully!"})

@app.route("/generate_schedule", methods=["POST"])
def generate_schedule():
    data = request.get_json()
    class_id = data["class_id"]
    day_of_week = data["day_of_week"]
    slot_order = data["slot_order"]  # list like ["period", "period", "break", "period"]

    created_entries = []
    for index, slot_type in enumerate(slot_order):
        new_entry = TimetableEntry(
            class_id=class_id,
            day_of_week=day_of_week,
            period_no=index + 1,
            entry_type=slot_type
        )
        db.session.add(new_entry)
        db.session.flush()  # get entry_id before commit
        created_entries.append(new_entry.entry_id)

    db.session.commit()
    return jsonify({"message": "Schedule slots created!", "entry_ids": created_entries})

@app.route("/fill_period_slot", methods=["POST"])
def fill_period_slot():
    data = request.get_json()
    entry = TimetableEntry.query.get(data["entry_id"])
    if not entry:
        return jsonify({"error": "Entry not found"}), 404

    # Reuse an existing course with the same name in this class, if one exists
    existing_course = Course.query.filter_by(
        class_id=entry.class_id,
        course_name=data["course_name"].strip()
    ).first()

    if existing_course:
        entry.course_id = existing_course.course_id
    else:
        new_course = Course(
            class_id=entry.class_id,
            course_name=data["course_name"].strip(),
            course_code=data.get("course_code", "")
        )
        db.session.add(new_course)
        db.session.flush()
        entry.course_id = new_course.course_id

    entry.start_time = data["start_time"]
    entry.end_time = data["end_time"]

    db.session.commit()
    return jsonify({"message": "Period filled successfully!"})
        
@app.route("/fill_break_slot", methods=["POST"])
def fill_break_slot():
    data = request.get_json()
    entry = TimetableEntry.query.get(data["entry_id"])
    if not entry:
        return jsonify({"error": "Entry not found"}), 404

    entry.label = data["label"]
    entry.start_time = data["start_time"]
    entry.end_time = data["end_time"]

    db.session.commit()
    return jsonify({"message": "Break filled successfully!"})

@app.route("/create_college_and_admin", methods=["POST"])
def create_college_and_admin():
    data = request.get_json()

    # Check college code isn't already taken
    existing_college = College.query.filter_by(college_code=data["college_code"].strip()).first()
    if existing_college:
        return jsonify({"error": "This college code is already taken. Choose another."}), 400

    existing_admin = Admin.query.filter_by(email=data["admin_email"].strip()).first()
    if existing_admin:
        return jsonify({"error": "This email is already registered."}), 400

    # Create the college
    new_college = College(
        college_name=data["college_name"].strip(),
        college_code=data["college_code"].strip()
    )
    db.session.add(new_college)
    db.session.flush()  # get college_id before commit

    # Create the admin account for this college
    hashed_pw = bcrypt.generate_password_hash(data["admin_password"]).decode('utf-8')
    new_admin = Admin(
        college_id=new_college.college_id,
        name=data["admin_name"].strip(),
        email=data["admin_email"].strip(),
        password_hash=hashed_pw,
        department_name=data.get("department_name", "").strip()
    )
    db.session.add(new_admin)
    db.session.commit()

    return jsonify({
        "message": "College and Admin account created successfully!",
        "college_id": new_college.college_id,
        "admin_id": new_admin.admin_id,
        "admin_name": new_admin.name
    })

@app.route("/promote_to_admin", methods=["POST"])
def promote_to_admin():
    data = request.get_json()

    faculty = Faculty.query.get(data["faculty_id"])
    if not faculty:
        return jsonify({"error": "Faculty not found"}), 404

    existing_admin = Admin.query.filter_by(email=faculty.email).first()
    if existing_admin:
        return jsonify({"error": "This person is already an admin"}), 400

    new_admin = Admin(
        college_id=faculty.college_id,
        name=faculty.name,
        email=faculty.email,
        password_hash=faculty.password_hash,  # reuse existing password
        department_name=faculty.subject_expertise or ""
    )
    db.session.add(new_admin)
    db.session.commit()

    return jsonify({"message": f"{faculty.name} is now an Admin!"})

@app.route("/faculty_list/<int:college_id>", methods=["GET"])
def get_faculty_list(college_id):
    faculty = Faculty.query.filter_by(college_id=college_id).all()
    result = []
    for f in faculty:
        is_admin = Admin.query.filter_by(email=f.email).first() is not None
        result.append({
            "faculty_id": f.faculty_id,
            "name": f.name,
            "email": f.email,
            "is_admin": is_admin
        })
    return jsonify(result)

@app.route("/admin_create_faculty", methods=["POST"])
def admin_create_faculty():
    data = request.get_json()

    admin = Admin.query.get(data["admin_id"])
    if not admin:
        return jsonify({"error": "Invalid admin"}), 400

    existing_faculty = Faculty.query.filter_by(email=data["email"].strip()).first()
    if existing_faculty:
        return jsonify({"error": f"{data['email']} is already registered"}), 400

    hashed_pw = bcrypt.generate_password_hash(data["password"]).decode('utf-8')

    new_faculty = Faculty(
        college_id=admin.college_id,
        name=data["name"].strip(),
        email=data["email"].strip(),
        password_hash=hashed_pw,
        subject_expertise=data.get("subject_expertise", "").strip()
    )
    db.session.add(new_faculty)
    db.session.commit()

    return jsonify({
        "message": f"{new_faculty.name} added successfully!",
        "faculty_id": new_faculty.faculty_id
    })

@app.route("/update_period_slot", methods=["POST"])
def update_period_slot():
    data = request.get_json()
    entry = TimetableEntry.query.get(data["entry_id"])
    if not entry or not entry.course_id:
        return jsonify({"error": "Entry not found"}), 404

    course = Course.query.get(entry.course_id)
    course.course_name = data["course_name"]
    course.course_code = data.get("course_code", "")
    entry.start_time = data["start_time"]
    entry.end_time = data["end_time"]

    db.session.commit()
    return jsonify({"message": "Period updated successfully!"})


@app.route("/update_break_slot", methods=["POST"])
def update_break_slot():
    data = request.get_json()
    entry = TimetableEntry.query.get(data["entry_id"])
    if not entry:
        return jsonify({"error": "Entry not found"}), 404

    entry.label = data["label"]
    entry.start_time = data["start_time"]
    entry.end_time = data["end_time"]

    db.session.commit()
    return jsonify({"message": "Break updated successfully!"})


@app.route("/delete_timetable_entry", methods=["POST"])
def delete_timetable_entry():
    data = request.get_json()
    entry = TimetableEntry.query.get(data["entry_id"])
    if not entry:
        return jsonify({"error": "Entry not found"}), 404

    course_id = entry.course_id
    delete_timetable_entries_cascade(entry_ids=[entry.entry_id])
    db.session.flush()

    if course_id:
        remaining_uses = TimetableEntry.query.filter_by(course_id=course_id).count()
        if remaining_uses == 0:
            course = Course.query.get(course_id)
            if course:
                db.session.delete(course)

    db.session.commit()
    return jsonify({"message": "Slot deleted successfully!"})


@app.route("/delete_day_schedule", methods=["POST"])
def delete_day_schedule():
    data = request.get_json()
    class_id = data["class_id"]
    day_of_week = data["day_of_week"]

    entries = TimetableEntry.query.filter_by(class_id=class_id, day_of_week=day_of_week).all()
    course_ids_to_check = set()
    entry_ids = [entry.entry_id for entry in entries]

    for entry in entries:
        if entry.course_id:
            course_ids_to_check.add(entry.course_id)

    if entry_ids:
        delete_timetable_entries_cascade(entry_ids=entry_ids)

    db.session.flush()

    for course_id in course_ids_to_check:
        remaining_uses = TimetableEntry.query.filter_by(course_id=course_id).count()
        if remaining_uses == 0:
            course = Course.query.get(course_id)
            if course:
                db.session.delete(course)

    db.session.commit()
    return jsonify({"message": f"{day_of_week} schedule deleted successfully!"})
    
@app.route("/notifications/<string:recipient_type>/<int:recipient_id>", methods=["GET"])
def get_notifications(recipient_type, recipient_id):
    notifs = UserNotification.query.filter_by(
        recipient_type=recipient_type,
        recipient_id=recipient_id
    ).order_by(UserNotification.created_at.desc()).all()

    result = []
    for n in notifs:
        result.append({
            "notification_id": n.notification_id,
            "message": n.message,
            "notif_type": n.notif_type,
            "reference_id": n.reference_id,
            "is_read": n.is_read,
            "created_at": n.created_at.strftime("%Y-%m-%d %H:%M") if n.created_at else ""
        })
    return jsonify(result)


@app.route("/unread_notification_count/<string:recipient_type>/<int:recipient_id>", methods=["GET"])
def get_unread_count(recipient_type, recipient_id):
    count = UserNotification.query.filter_by(
        recipient_type=recipient_type,
        recipient_id=recipient_id,
        is_read=False
    ).count()
    return jsonify({"count": count})



@app.route("/mark_notification_read", methods=["POST"])
def mark_notification_read():
    data = request.get_json()
    notif = UserNotification.query.get(data["notification_id"])
    if notif:
        notif.is_read = True
        db.session.commit()
    return jsonify({"message": "Marked as read"})

@app.route("/mark_all_notifications_read", methods=["POST"])
def mark_all_notifications_read():
    data = request.get_json()
    recipient_type = data.get("recipient_type")
    recipient_id = data.get("recipient_id")
    if recipient_type and recipient_id:
        UserNotification.query.filter_by(
            recipient_type=recipient_type,
            recipient_id=recipient_id,
            is_read=False
        ).update({UserNotification.is_read: True})
        db.session.commit()
    return jsonify({"message": "All marked as read"})

@app.route("/course_detail/<int:course_id>", methods=["GET"])
def get_course_detail(course_id):
    course = Course.query.get(course_id)
    if not course:
        return jsonify({"error": "Course not found"}), 404

    faculty_info = None
    if course.faculty_id:
        faculty = Faculty.query.get(course.faculty_id)
        faculty_info = {
            "name": faculty.name,
            "email": faculty.email,
            "subject_expertise": faculty.subject_expertise or "Not specified"
        }

    return jsonify({
        "course_name": course.course_name,
        "course_code": course.course_code or "N/A",
        "faculty": faculty_info
    })

@app.route("/cc_details/<int:class_id>", methods=["GET"])
def get_cc_details(class_id):
    class_obj = Class.query.get(class_id)
    if not class_obj or not class_obj.cc_faculty_id:
        return jsonify({"error": "No CC assigned"}), 404

    faculty = Faculty.query.get(class_obj.cc_faculty_id)
    return jsonify({
        "faculty_id": faculty.faculty_id,
        "name": faculty.name,
        "email": faculty.email,
        "subject_expertise": faculty.subject_expertise or "Not specified"
    })

@app.route("/remove_cc", methods=["POST"])
def remove_cc():
    data = request.get_json()
    class_obj = Class.query.get(data["class_id"])
    if not class_obj:
        return jsonify({"error": "Class not found"}), 404

    class_obj.cc_faculty_id = None
    db.session.commit()
    return jsonify({"message": "CC removed successfully!"})

@app.route("/class_updates/<int:class_id>", methods=["GET"])
def get_class_updates(class_id):
    class_obj = Class.query.get(class_id)
    if not class_obj:
        return jsonify([])

    entries = TimetableEntry.query.filter_by(class_id=class_id).all()
    entry_ids = [e.entry_id for e in entries]

    updates = []

    # 1. Upcoming and recent College Holidays
    holidays = Holiday.query.filter_by(college_id=class_obj.college_id).order_by(Holiday.holiday_date.desc()).limit(5).all()
    for h in holidays:
        updates.append({
            "type": "holiday",
            "title": f"Holiday Notice ({h.holiday_date})",
            "message": f"{h.reason or 'College Holiday'} on {h.holiday_date}.",
            "created_at": f"{h.holiday_date} 00:00"
        })

    # 2. Confirmed cover requests (substitute assigned)
    if entry_ids:
        confirmed_leaves = Leave.query.filter(
            Leave.entry_id.in_(entry_ids),
            Leave.status == "confirmed"
        ).order_by(Leave.created_at.desc()).limit(10).all()

        for leave in confirmed_leaves:
            entry = TimetableEntry.query.get(leave.entry_id)
            original = Faculty.query.get(leave.faculty_id)
            substitute = Faculty.query.get(leave.confirmed_faculty_id) if leave.confirmed_faculty_id else None
            course = Course.query.get(entry.course_id) if entry and entry.course_id else None

            if substitute and course:
                updates.append({
                    "type": "substitute_confirmed",
                    "title": "Substitute Assigned",
                    "message": f"{original.name} is on leave for {course.course_name} ({entry.day_of_week} Period {entry.period_no}). {substitute.name} will take the class.",
                    "created_at": leave.created_at.strftime("%Y-%m-%d %H:%M") if leave.created_at else ""
                })

        # 3. Open leaves with no substitute yet (free period)
        open_leaves = Leave.query.filter(
            Leave.entry_id.in_(entry_ids),
            Leave.status.in_(["open", "pending_requests"])
        ).order_by(Leave.created_at.desc()).limit(10).all()

        for leave in open_leaves:
            entry = TimetableEntry.query.get(leave.entry_id)
            original = Faculty.query.get(leave.faculty_id)
            course = Course.query.get(entry.course_id) if entry and entry.course_id else None

            if course and original:
                updates.append({
                    "type": "free_period",
                    "title": "Faculty on Leave",
                    "message": f"{original.name} is on leave for {course.course_name} ({entry.day_of_week} Period {entry.period_no}). Free period or substitute pending.",
                    "created_at": leave.created_at.strftime("%Y-%m-%d %H:%M") if leave.created_at else ""
                })

        # 4. Swapped periods
        swaps = SwapRequest.query.filter(
            db.or_(
                SwapRequest.requester_entry_id.in_(entry_ids),
                SwapRequest.target_entry_id.in_(entry_ids)
            ),
            SwapRequest.status == "accepted"
        ).order_by(SwapRequest.resolved_at.desc()).limit(10).all()

        for swap in swaps:
            requester = Faculty.query.get(swap.requester_faculty_id)
            target = Faculty.query.get(swap.target_faculty_id)
            updates.append({
                "type": "swap",
                "title": "Class Swapped",
                "message": f"Period swap confirmed between {requester.name} and {target.name}.",
                "created_at": swap.resolved_at.strftime("%Y-%m-%d %H:%M") if swap.resolved_at else ""
            })

    # Sort all combined updates by time, most recent first
    updates.sort(key=lambda x: x.get("created_at", ""), reverse=True)

    return jsonify(updates[:15])

@app.route("/faculty_current_cc/<int:faculty_id>", methods=["GET"])
def get_faculty_current_cc(faculty_id):
    class_obj = Class.query.filter_by(cc_faculty_id=faculty_id).first()
    if not class_obj:
        return jsonify({"has_cc": False})
    return jsonify({
        "has_cc": True,
        "class_id": class_obj.class_id,
        "year": class_obj.year,
        "section": class_obj.section
    })

@app.route("/cc_request_detail/<int:cc_request_id>", methods=["GET"])
def get_cc_request_detail(cc_request_id):
    cc_request = CCRequest.query.get(cc_request_id)
    if not cc_request:
        return jsonify({"error": "Request not found"}), 404
    class_obj = Class.query.get(cc_request.class_id)
    return jsonify({
        "cc_request_id": cc_request.cc_request_id,
        "class_id": class_obj.class_id,
        "year": class_obj.year,
        "section": class_obj.section,
        "status": cc_request.status
    })

@app.route("/classes_by_semester/<int:semester_id>", methods=["GET"])
def get_classes_by_semester(semester_id):
    classes = Class.query.filter_by(semester_id=semester_id).all()
    result = []
    for c in classes:
        result.append({
            "class_id": c.class_id,
            "year": c.year,
            "section": c.section,
            "department": c.department,
            "cc_faculty_id": c.cc_faculty_id
        })
    return jsonify(result)

@app.route("/admin_faculty_context/<int:admin_id>", methods=["GET"])
def get_admin_faculty_context(admin_id):
    admin = Admin.query.get(admin_id)
    if not admin:
        return jsonify({"error": "Admin not found"}), 404

    faculty = Faculty.query.filter_by(email=admin.email).first()
    if not faculty:
        return jsonify({"has_faculty_record": False})

    return jsonify({
        "has_faculty_record": True,
        "faculty_id": faculty.faculty_id,
        "faculty_name": faculty.name
    })

@app.route("/admin_self_assign_cc", methods=["POST"])
def admin_self_assign_cc():
    data = request.get_json()
    class_id = data["class_id"]
    faculty_id = data["faculty_id"]

    class_obj = Class.query.get(class_id)
    if not class_obj:
        return jsonify({"error": "Class not found"}), 404

    faculty = Faculty.query.get(faculty_id)
    if not faculty:
        return jsonify({"error": "Faculty record not found"}), 404

    # Clear this faculty from any OTHER class where they're currently CC,
    # same rule the normal accept-flow enforces (one CC class at a time).
    old_classes = Class.query.filter(
        Class.cc_faculty_id == faculty_id,
        Class.class_id != class_id
    ).all()
    for old_class in old_classes:
        old_class.cc_faculty_id = None

    class_obj.cc_faculty_id = faculty_id
    db.session.commit()

    return jsonify({
        "message": f"You are now CC for {class_obj.year} - {class_obj.section}!"
    })

@app.route("/swap_responses/<int:faculty_id>", methods=["GET"])
def get_swap_responses(faculty_id):
    requests = SwapRequest.query.filter(
        SwapRequest.requester_faculty_id == faculty_id,
        SwapRequest.status != "pending"
    ).order_by(SwapRequest.resolved_at.desc()).all()

    result = []
    for r in requests:
        requester_entry = TimetableEntry.query.get(r.requester_entry_id)
        target_entry = TimetableEntry.query.get(r.target_entry_id)
        requester_class = Class.query.get(requester_entry.class_id)
        target_class = Class.query.get(target_entry.class_id)
        target_faculty = Faculty.query.get(r.target_faculty_id)
        requester_course = Course.query.get(requester_entry.course_id) if requester_entry.course_id else None
        target_course = Course.query.get(target_entry.course_id) if target_entry.course_id else None

        result.append({
            "swap_id": r.swap_id,
            "status": r.status,
            "target_name": target_faculty.name if target_faculty else "Unknown",
            "your_class": f"{requester_class.year} - {requester_class.department} - {requester_class.section}",
            "your_slot": f"{requester_entry.day_of_week} Period {requester_entry.period_no}",
            "your_course": requester_course.course_name if requester_course else "Unassigned",
            "their_class": f"{target_class.year} - {target_class.department} - {target_class.section}",
            "their_slot": f"{target_entry.day_of_week} Period {target_entry.period_no}",
            "their_course": target_course.course_name if target_course else "Unassigned",
            "rejection_reason": r.rejection_reason or "",
            "resolved_at": r.resolved_at.strftime("%Y-%m-%d %H:%M") if r.resolved_at else ""
        })
    return jsonify(result)


@app.route("/admin_self_assign_course_faculty", methods=["POST"])
def admin_self_assign_course_faculty():
    data = request.get_json()
    course_id = data["course_id"]
    faculty_id = data["faculty_id"]

    course = Course.query.get(course_id)
    if not course:
        return jsonify({"error": "Course not found"}), 404

    faculty = Faculty.query.get(faculty_id)
    if not faculty:
        return jsonify({"error": "Faculty record not found"}), 404

    course.faculty_id = faculty_id
    db.session.commit()

    return jsonify({
        "message": f"You are now teaching {course.course_name}!"
    })

@app.route("/update_class", methods=["POST"])
def update_class():
    data = request.get_json()
    class_obj = Class.query.get(data["class_id"])
    if not class_obj:
        return jsonify({"error": "Class not found"}), 404

    admin_id = data.get("admin_id")
    faculty_id = data.get("faculty_id")

    if admin_id is not None:
        admin = Admin.query.get(admin_id)
        if not admin or admin.college_id != class_obj.college_id:
            return jsonify({"error": "Not authorized to edit this class"}), 403
    elif faculty_id is not None:
        if class_obj.cc_faculty_id != faculty_id:
            return jsonify({"error": "Only this class's CC can edit it"}), 403
    else:
        return jsonify({"error": "Not authorized to edit this class"}), 403

    new_year = data["year"].strip()
    new_section = data["section"].strip()
    new_department = data["department"].strip()

    existing = Class.query.filter(
        Class.college_id == class_obj.college_id,
        Class.semester_id == class_obj.semester_id,
        Class.year == new_year,
        Class.section == new_section,
        Class.department == new_department,
        Class.class_id != class_obj.class_id
    ).first()
    if existing:
        return jsonify({"error": "A class with these details already exists in this semester"}), 400

    class_obj.year = new_year
    class_obj.section = new_section
    class_obj.department = new_department
    db.session.commit()

    return jsonify({"message": "Class updated successfully!"})

@app.route("/my_sent_swap_requests/<int:faculty_id>", methods=["GET"])
def get_my_sent_swap_requests(faculty_id):
    requests = SwapRequest.query.filter_by(
        requester_faculty_id=faculty_id
    ).order_by(SwapRequest.requested_at.desc()).all()

    result = []
    for r in requests:
        requester_entry = TimetableEntry.query.get(r.requester_entry_id)
        target_entry = TimetableEntry.query.get(r.target_entry_id)
        requester_class = Class.query.get(requester_entry.class_id)
        target_class = Class.query.get(target_entry.class_id)
        target_faculty = Faculty.query.get(r.target_faculty_id)
        requester_course = Course.query.get(requester_entry.course_id) if requester_entry.course_id else None
        target_course = Course.query.get(target_entry.course_id) if target_entry.course_id else None

        result.append({
            "swap_id": r.swap_id,
            "status": r.status,
            "target_name": target_faculty.name if target_faculty else "Unknown",
            "requester_class": f"{requester_class.year} - {requester_class.department} - {requester_class.section}",
            "requester_slot": f"{requester_entry.day_of_week} Period {requester_entry.period_no}",
            "requester_course": requester_course.course_name if requester_course else None,
            "target_class": f"{target_class.year} - {target_class.department} - {target_class.section}",
            "target_slot": f"{target_entry.day_of_week} Period {target_entry.period_no}",
            "target_course": target_course.course_name if target_course else None,
            "rejection_reason": r.rejection_reason,
        })
    return jsonify(result)

def cleanup_expired_statuses():
    with app.app_context():
        cutoff = datetime.now() - timedelta(days=7)
        cutoff_date_str = cutoff.strftime("%Y-%m-%d")

        # 1. Revert any leave (open, pending_requests, or confirmed) whose
        # leave_date is more than a week old — restores the original
        # faculty's period and clears any substitute assignment.
        old_leaves = Leave.query.filter(Leave.leave_date < cutoff_date_str).all()
        reverted_leaves = 0
        for leave in old_leaves:
            entry = TimetableEntry.query.get(leave.entry_id)
            class_obj = Class.query.get(entry.class_id) if entry else None

            if leave.status == "confirmed" and leave.confirmed_faculty_id and entry and class_obj:
                substitute = Faculty.query.get(leave.confirmed_faculty_id)
                if substitute:
                    create_notification(
                        "faculty", substitute.faculty_id,
                        f"You're no longer covering {class_obj.year} - {class_obj.department} - "
                        f"{class_obj.section}, {entry.day_of_week} Period {entry.period_no} — "
                        f"that week has passed and the schedule reset to normal.",
                        "cover_request"
                    )

            if entry:
                entry.status_color = "normal"

            CoverRequest.query.filter_by(leave_id=leave.leave_id).delete()
            db.session.delete(leave)
            reverted_leaves += 1

        # 2. Revert any accepted swap resolved more than a week ago — swaps
        # are self-inverse, so exchanging course_id back undoes it exactly.
        old_swaps = SwapRequest.query.filter(
            SwapRequest.status == "accepted",
            SwapRequest.resolved_at < cutoff
        ).all()
        reverted_swaps = 0
        for swap in old_swaps:
            requester_entry = TimetableEntry.query.get(swap.requester_entry_id)
            target_entry = TimetableEntry.query.get(swap.target_entry_id)
            if requester_entry and target_entry:
                requester_entry.course_id, target_entry.course_id = (
                    target_entry.course_id,
                    requester_entry.course_id,
                )
                requester_entry.status_color = "normal"
                target_entry.status_color = "normal"
            db.session.delete(swap)
            reverted_swaps += 1

        db.session.commit()
        print(
            f"[CLEANUP] Reverted {reverted_leaves} leave(s) and "
            f"{reverted_swaps} swap(s) older than 7 days."
        )

@app.route("/leave_by_entry/<int:entry_id>", methods=["GET"])
def get_leave_by_entry(entry_id):
    leave = Leave.query.filter_by(entry_id=entry_id).filter(Leave.status != "confirmed").first()
    if not leave:
        return jsonify({"error": "No active leave found for this entry"}), 404
    return jsonify({"leave_id": leave.leave_id})

@app.route("/invite_substitute", methods=["POST"])
def invite_substitute():
    data = request.get_json()
    leave = Leave.query.get(data["leave_id"])
    if not leave:
        return jsonify({"error": "Leave not found"}), 404
    if leave.status == "confirmed":
        return jsonify({"error": "This leave slot is already confirmed"}), 400

    faculty = Faculty.query.get(data["faculty_id"])
    if not faculty:
        return jsonify({"error": "Faculty not found"}), 404

    entry = TimetableEntry.query.get(leave.entry_id)
    class_obj = Class.query.get(entry.class_id)
    course = Course.query.get(entry.course_id) if entry.course_id else None

    # Check if there is already a CoverRequest for this faculty on this leave
    existing_req = CoverRequest.query.filter_by(
        leave_id=leave.leave_id,
        requesting_faculty_id=faculty.faculty_id
    ).first()

    if not existing_req:
        existing_req = CoverRequest(
            leave_id=leave.leave_id,
            requesting_faculty_id=faculty.faculty_id,
            status="invited"
        )
        db.session.add(existing_req)
        db.session.commit()
    else:
        existing_req.status = "invited"
        db.session.commit()

    leave.status = "pending_requests"
    db.session.commit()

    original_faculty = Faculty.query.get(leave.faculty_id)
    create_notification(
        "faculty", faculty.faculty_id,
        f"{original_faculty.name} invited you to cover "
        f"{class_obj.year} - {class_obj.department} - {class_obj.section}, "
        f"{entry.day_of_week} Period {entry.period_no}"
        f"{f' ({course.course_name})' if course else ''}.",
        "substitute_invite", reference_id=existing_req.cover_req_id
    )

    return jsonify({"message": f"Invitation request sent to {faculty.name}!"})

@app.route("/resolve_substitute_invite", methods=["POST"])
def resolve_substitute_invite():
    data = request.get_json()
    cover_req_id = data["cover_req_id"]
    decision = data["decision"] # 'accepted' or 'rejected'

    cover_req = CoverRequest.query.get(cover_req_id)
    if not cover_req:
        return jsonify({"error": "Request not found"}), 404

    leave = Leave.query.get(cover_req.leave_id)
    if not leave:
        return jsonify({"error": "Leave not found"}), 404

    entry = TimetableEntry.query.get(leave.entry_id)
    class_obj = Class.query.get(entry.class_id)
    course = Course.query.get(entry.course_id) if entry.course_id else None

    substitute = Faculty.query.get(cover_req.requesting_faculty_id)
    original_faculty = Faculty.query.get(leave.faculty_id)

    if decision == "accepted":
        if leave.status == "confirmed":
            return jsonify({"error": "This leave is already confirmed"}), 400

        cover_req.status = "accepted"

        # Reject other cover requests for this leave
        other_requests = CoverRequest.query.filter(
            CoverRequest.leave_id == leave.leave_id,
            CoverRequest.cover_req_id != cover_req.cover_req_id
        ).all()
        for r in other_requests:
            r.status = "rejected"
            if r.status == "pending":
                vol = Faculty.query.get(r.requesting_faculty_id)
                if vol:
                    create_notification(
                        "faculty", vol.faculty_id,
                        f"The {class_obj.year} - {class_obj.department} - {class_obj.section}, "
                        f"{entry.day_of_week} Period {entry.period_no} slot was assigned to someone else.",
                        "cover_request"
                    )

        leave.status = "confirmed"
        leave.confirmed_faculty_id = substitute.faculty_id
        leave.confirmed_by_role = "faculty"

        entry.status_color = "confirmed_cover"

        db.session.commit()

        # Send confirmation notification to TOC faculty (original absent faculty)
        create_notification(
            "faculty", original_faculty.faculty_id,
            f"{substitute.name} accepted your invite to cover "
            f"{class_obj.year} - {class_obj.department} - {class_obj.section}, "
            f"{entry.day_of_week} Period {entry.period_no}"
            f"{f' ({course.course_name})' if course else ''}.",
            "cover_confirmed"
        )
        create_notification(
            "class", entry.class_id,
            f"Timetable change: {substitute.name} will substitute for {original_faculty.name} on {entry.day_of_week} Period {entry.period_no}",
            "cover_confirmed"
        )
        return jsonify({"message": "Invitation accepted successfully!"})

    else: # rejected
        cover_req.status = "rejected"
        db.session.commit()

        # Send decline notification to original faculty
        create_notification(
            "faculty", original_faculty.faculty_id,
            f"{substitute.name} declined your invite to cover "
            f"{class_obj.year} - {class_obj.department} - {class_obj.section}, "
            f"{entry.day_of_week} Period {entry.period_no}"
            f"{f' ({course.course_name})' if course else ''}.",
            "cover_declined"
        )
        return jsonify({"message": "Invitation declined."})

@app.route("/class_info/<int:class_id>", methods=["GET"])
def get_class_info(class_id):
    class_obj = Class.query.get(class_id)
    if not class_obj:
        return jsonify({"error": "Class not found"}), 404

    cc_name = None
    if class_obj.cc_faculty_id:
        cc = Faculty.query.get(class_obj.cc_faculty_id)
        cc_name = cc.name if cc else None

    return jsonify({
        "class_id": class_obj.class_id,
        "year": class_obj.year,
        "section": class_obj.section,
        "department": class_obj.department,
        "room_number": class_obj.room_number,
        "cc_name": cc_name  # always derived automatically — never manually typed
    })


@app.route("/update_class_room", methods=["POST"])
def update_class_room():
    data = request.get_json()
    class_obj = Class.query.get(data["class_id"])
    if not class_obj:
        return jsonify({"error": "Class not found"}), 404

    class_obj.room_number = data.get("room_number", "").strip() or None
    db.session.commit()

    return jsonify({"message": "Room number updated!"})

@app.route("/mark_holiday", methods=["POST"])
def mark_holiday():
    data = request.get_json()
    college_id = data.get("college_id")
    if not college_id and data.get("class_id"):
        class_obj = Class.query.get(data["class_id"])
        if class_obj:
            college_id = class_obj.college_id
    if not college_id:
        return jsonify({"error": "college_id or class_id required"}), 400

    existing = Holiday.query.filter_by(
        college_id=college_id, holiday_date=data["holiday_date"]
    ).first()
    if existing:
        return jsonify({"error": "This date is already marked as a holiday"}), 400
    new_holiday = Holiday(
        college_id=college_id,
        holiday_date=data["holiday_date"],
        reason=data.get("reason", "Declared Holiday")
    )
    db.session.add(new_holiday)
    db.session.commit()
    return jsonify({"message": "Holiday marked!", "holiday_id": new_holiday.holiday_id})

@app.route("/holidays/<int:college_id>", methods=["GET"])
def get_holidays(college_id):
    holidays = Holiday.query.filter_by(college_id=college_id).order_by(Holiday.holiday_date.desc()).all()
    result = []
    for h in holidays:
        result.append({
            "holiday_id": h.holiday_id,
            "holiday_date": h.holiday_date,
            "reason": h.reason or ""
        })
    return jsonify(result)

@app.route("/delete_holiday", methods=["POST"])
def delete_holiday():
    data = request.get_json()
    holiday = Holiday.query.get(data["holiday_id"])
    if not holiday:
        return jsonify({"error": "Holiday not found"}), 404
    db.session.delete(holiday)
    db.session.commit()
    return jsonify({"message": "Holiday removed"})

@app.route("/holidays_for_class_week/<int:class_id>", methods=["GET"])
def get_holidays_for_class_week(class_id):
    class_obj = Class.query.get(class_id)
    if not class_obj:
        return jsonify({})
    holidays = Holiday.query.filter_by(college_id=class_obj.college_id).all()
    return jsonify({
        h.holiday_date: {"holiday_id": h.holiday_id, "reason": h.reason or ""}
        for h in holidays
    })


if os.environ.get('WERKZEUG_RUN_MAIN') == 'true' or not app.debug:
    scheduler = BackgroundScheduler()
    scheduler.add_job(check_pending_leaves, 'interval', minutes=1)
    scheduler.add_job(cleanup_expired_statuses, 'interval', hours=6)
    scheduler.add_job(reset_weekly_swaps, CronTrigger(day_of_week='mon', hour=0, minute=0))
    scheduler.add_job(reset_weekly_leaves, CronTrigger(day_of_week='mon', hour=0, minute=1))
    scheduler.start()

if __name__ == "__main__":
    app.run(debug=True)